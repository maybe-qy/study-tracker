#!/usr/bin/env python3
"""Record one exam: append to 成绩总表.xlsx and create immutable .md archive.

Usage:
  echo '{"workspace": "...", "exam_name": "...", ...}' | python record_exam.py
  python record_exam.py --workspace "d:/研究/学升" < data.json

Input JSON fields (required):
  workspace, exam_name, exam_date, exam_type, grade, total_score

Input JSON fields (optional):
  cn_score, math_score, en_score,
  sub1_name, sub1_raw, sub1_assigned, sub1_confidence,
  sub2_name, sub2_raw, sub2_assigned, sub2_confidence,
  sub3_name, sub3_raw, sub3_assigned, sub3_confidence,
  city_rank, city_total, alliance_rank, alliance_total,
  school_rank, school_total, special_line, unexamined_top_students, excellent_line, notes
"""

import json
import os
import sys

from openpyxl import load_workbook

REQUIRED = ["workspace", "exam_name", "exam_date", "exam_type", "grade", "total_score"]
OPTIONAL = [
    "cn_score", "math_score", "en_score",
    "sub1_name", "sub1_raw", "sub1_assigned", "sub1_confidence",
    "sub2_name", "sub2_raw", "sub2_assigned", "sub2_confidence",
    "sub3_name", "sub3_raw", "sub3_assigned", "sub3_confidence",
    "city_rank", "city_total", "alliance_rank", "alliance_total",
    "school_rank", "school_total", "special_line", "unexamined_top_students",
    "excellent_line", "score_scale", "school_type", "rank_type", "notes",
]

EXCEL_COLS = [
    "考试名", "日期", "类型", "年级",
    "语文", "数学", "英语",
    "选科1名称", "选科1原始分", "选科1赋分", "选科1赋分置信度",
    "选科2名称", "选科2原始分", "选科2赋分", "选科2赋分置信度",
    "选科3名称", "选科3原始分", "选科3赋分", "选科3赋分置信度",
    "总分",
    "市/联盟排名", "市/联盟总人数",
    "校排名", "校总人数",
    "特控线", "重点班未参考人数", "优划线", "满分制", "学校类型", "排名类型", "备注",
]

MD_TEMPLATE = """# {exam_name}

- **日期**：{exam_date}
- **类型**：{exam_type}
- **年级**：{grade}

## 各科成绩

| 科目 | 原始分 | 赋分 | 赋分置信度 |
|------|--------|------|-----------|
| 语文 | {cn_score} | - | - |
| 数学 | {math_score} | - | - |
| 英语 | {en_score} | - | - |
| {sub1_name} | {sub1_raw} | {sub1_assigned} | {sub1_confidence} |
| {sub2_name} | {sub2_raw} | {sub2_assigned} | {sub2_confidence} |
| {sub3_name} | {sub3_raw} | {sub3_assigned} | {sub3_confidence} |

- **总分**：{total_score}
- **总分校验**：{check_note}

## 排名信息

- 市/联盟排名：{city_alliance_rank}
- 市/联盟总人数：{city_alliance_total}
- 校排名：{school_rank}
- 校总人数：{school_total}

## 分数线

- 特控线：{special_line}
- 优划线：{excellent_line}

## 备注

{notes}
"""


def validate(data):
    missing = [f for f in REQUIRED if f not in data or data[f] is None or (isinstance(data[f], str) and not data[f].strip())]
    if missing:
        return f"缺少必填字段: {', '.join(missing)}"
    return None


def build_row(data):
    """Map input JSON to Excel column order."""
    # Combine city/alliance rank into one field
    city_alliance_rank = data.get("city_rank") or data.get("alliance_rank") or None
    city_alliance_total = data.get("city_total") or data.get("alliance_total") or None

    return [
        data["exam_name"],
        data["exam_date"],
        data["exam_type"],
        data["grade"],
        data.get("cn_score"),
        data.get("math_score"),
        data.get("en_score"),
        data.get("sub1_name"),
        data.get("sub1_raw"),
        data.get("sub1_assigned"),
        data.get("sub1_confidence"),
        data.get("sub2_name"),
        data.get("sub2_raw"),
        data.get("sub2_assigned"),
        data.get("sub2_confidence"),
        data.get("sub3_name"),
        data.get("sub3_raw"),
        data.get("sub3_assigned"),
        data.get("sub3_confidence"),
        data["total_score"],
        city_alliance_rank,
        city_alliance_total,
        data.get("school_rank"),
        data.get("school_total"),
        data.get("special_line"),
        data.get("unexamined_top_students"),
        data.get("excellent_line"),
        data.get("score_scale") or 750,
        data.get("school_type"),
        data.get("rank_type"),
        data.get("notes"),
    ]


def create_md(workspace, data):
    """Write immutable markdown record to 个体数据/."""
    # Calculate sum check: 语数英原始分 + 选科赋分（有赋分用赋分，没有用原始分）
    # 450分制考试总分仅含语数英，不包含选科
    score_scale = int(data.get("score_scale", 750) or 750)
    if score_scale <= 0:
        score_scale = 750

    def _safe_float(v):
        try:
            return float(v) if v is not None else 0
        except (ValueError, TypeError):
            return 0
    cn = data.get("cn_score")
    math = data.get("math_score")
    en = data.get("en_score")
    comparison_sum = _safe_float(cn) + _safe_float(math) + _safe_float(en)
    if score_scale != 450:
        for i in range(1, 4):
            assigned = data.get(f"sub{i}_assigned")
            raw = data.get(f"sub{i}_raw")
            if assigned is not None and assigned != "":
                comparison_sum += _safe_float(assigned)
            elif raw is not None and raw != "":
                comparison_sum += _safe_float(raw)
    total = data["total_score"]
    if abs(comparison_sum - _safe_float(total)) > 0.5:
        check_note = f"各科加总（选科有赋分用赋分）= {comparison_sum}，≠ 总分 {total}（用户确认以总分为准）"
    else:
        check_note = f"各科加总 = {comparison_sum}，与总分一致"

    # Format rank info
    city_alliance_rank = data.get("city_rank") or data.get("alliance_rank") or "-"
    city_alliance_total = data.get("city_total") or data.get("alliance_total") or "-"
    school_rank = data.get("school_rank") or "-"
    school_total = data.get("school_total") or "-"

    def fmt(v):
        """Format value for MD: 0 is valid, only None/empty shows as '-'."""
        return str(v) if v is not None and v != "" else "-"

    content = MD_TEMPLATE.format(
        exam_name=data["exam_name"],
        exam_date=data["exam_date"],
        exam_type=data["exam_type"],
        grade=data["grade"],
        cn_score=fmt(data.get("cn_score")),
        math_score=fmt(data.get("math_score")),
        en_score=fmt(data.get("en_score")),
        sub1_name=fmt(data.get("sub1_name")),
        sub1_raw=fmt(data.get("sub1_raw")),
        sub1_assigned=fmt(data.get("sub1_assigned")),
        sub1_confidence=fmt(data.get("sub1_confidence")),
        sub2_name=fmt(data.get("sub2_name")),
        sub2_raw=fmt(data.get("sub2_raw")),
        sub2_assigned=fmt(data.get("sub2_assigned")),
        sub2_confidence=fmt(data.get("sub2_confidence")),
        sub3_name=fmt(data.get("sub3_name")),
        sub3_raw=fmt(data.get("sub3_raw")),
        sub3_assigned=fmt(data.get("sub3_assigned")),
        sub3_confidence=fmt(data.get("sub3_confidence")),
        total_score=total,
        check_note=check_note,
        city_alliance_rank=city_alliance_rank,
        city_alliance_total=city_alliance_total,
        school_rank=school_rank,
        school_total=school_total,
        special_line=fmt(data.get("special_line")),
        excellent_line=fmt(data.get("excellent_line")),
        notes=data.get("notes") or "",
    )

    safe_name = f"{data['exam_date']}_{data['exam_name']}"
    for ch in '/\\:*?"<>|':
        safe_name = safe_name.replace(ch, "_")
    md_dir = os.path.join(workspace, "data", "personal", "individual")
    os.makedirs(md_dir, exist_ok=True)
    md_path = os.path.join(md_dir, f"{safe_name}.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(content)
    return md_path


def update_subject_tracking(workspace, data):
    """Append one row per subject to 单科追踪.xlsx."""
    tracking_path = os.path.join(workspace, "data", "personal", "单科追踪.xlsx")
    if not os.path.exists(tracking_path):
        return

    wb = None
    try:
        wb = load_workbook(tracking_path)

        # Map subject names to sheet names
        subjects = [
            ("语文追踪", "语文", data.get("cn_score"), None, "B"),
            ("数学追踪", "数学", data.get("math_score"), None, "B"),
            ("英语追踪", "英语", data.get("en_score"), None, "B"),
        ]

        # Determine 选科 names — use the same logic for consistency
        for i in range(1, 4):
            sub_name = data.get(f"sub{i}_name")
            sub_raw = data.get(f"sub{i}_raw")
            sub_assigned = data.get(f"sub{i}_assigned")
            sub_conf = data.get(f"sub{i}_confidence") or "B"
            sheet_name = f"选科{i}追踪"
            subjects.append((sheet_name, sub_name, sub_raw, sub_assigned, sub_conf))

        for sheet_name, subject_name, raw_val, assigned_val, conf_val in subjects:
            if raw_val is None or raw_val == "":
                continue
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            ws.append([
                data.get("exam_name", ""),
                data.get("exam_date", ""),
                raw_val,
                assigned_val if assigned_val is not None and assigned_val != "" else "",
                conf_val if conf_val is not None and conf_val != "" else "",
            ])

        wb.save(tracking_path)
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _check_duplicate(ws, exam_name, exam_date):
    """Check if a record with the same exam_name and exam_date already exists."""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2:
            continue
        existing_name = str(row[0]).strip() if row[0] else ""
        existing_date = str(row[1]).strip() if row[1] else ""
        if existing_name == str(exam_name).strip() and existing_date == str(exam_date).strip():
            return True
    return False


def run(data):
    workspace = os.path.abspath(data["workspace"])
    excel_path = os.path.join(workspace, "data", "personal", "成绩总表.xlsx")

    if not os.path.exists(excel_path):
        return {"status": "error", "reason": "成绩总表.xlsx 不存在，请先运行 setup_workspace.py"}

    try:
        wb = load_workbook(excel_path)
    except Exception as e:
        return {"status": "error", "reason": f"成绩总表.xlsx 读取失败: {e}"}

    if "成绩总表" not in wb.sheetnames:
        wb.close()
        return {"status": "error", "reason": "成绩总表.xlsx 中缺少「成绩总表」Sheet"}

    ws = wb["成绩总表"]

    # 重复录入检测
    exam_name = data.get("exam_name", "")
    exam_date = data.get("exam_date", "")
    if _check_duplicate(ws, exam_name, exam_date):
        wb.close()
        return {
            "status": "error",
            "reason": f"已存在相同考试记录（{exam_name} / {exam_date}），请勿重复录入。如需修改请直接编辑 Excel 文件。",
        }

    row = build_row(data)

    saved_max_row = None
    try:
        ws.append(row)
        saved_max_row = ws.max_row
        wb.save(excel_path)
    except Exception as e:
        wb.close()
        return {"status": "error", "reason": f"成绩总表写入失败: {e}"}
    finally:
        try:
            wb.close()
        except Exception:
            pass

    # 更新单科追踪（失败不影响主流程，但会返回警告）
    warnings = []
    try:
        update_subject_tracking(workspace, data)
    except Exception as e:
        warnings.append(f"单科追踪更新失败: {e}")

    # 生成 MD 存档（失败不影响主流程，但会返回警告）
    md_path = None
    try:
        md_path = create_md(workspace, data)
    except Exception as e:
        warnings.append(f"MD存档生成失败: {e}")

    result = {
        "status": "ok",
        "row": saved_max_row,
        "record_index": (saved_max_row - 1) if saved_max_row else 0,  # 第几条记录（排除表头行）
        "md_path": md_path,
    }
    if warnings:
        result["warnings"] = warnings
    return result


def main():
    try:
        data = json.loads(sys.stdin.read())
    except json.JSONDecodeError as e:
        print(json.dumps({"status": "error", "reason": f"JSON 解析失败: {e}"}, ensure_ascii=False))
        sys.exit(1)

    err = validate(data)
    if err:
        print(json.dumps({"status": "error", "reason": err}, ensure_ascii=False))
        sys.exit(1)

    try:
        result = run(data)
    except Exception as e:
        result = {"status": "error", "reason": f"未知错误: {e}"}

    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
