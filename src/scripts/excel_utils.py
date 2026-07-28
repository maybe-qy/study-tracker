#!/usr/bin/env python3
"""Shared Excel utilities for reading worksheets with title-row detection.

Used by calc_equivalent.py and generate_reports.py to eliminate code duplication.
"""

import os

from openpyxl import load_workbook


# 已知列名关键词集合，用于检测第1行是标题还是表头
KNOWN_COLUMN_KEYWORDS = {
    "分数", "排名", "累计人数", "年份", "人数", "特控线", "上线",
    "下限", "上限", "科目", "得分", "等效分", "原始分", "赋分",
    "总分", "成绩", "名称", "考试名", "日期", "置信度", "方法",
    "score", "rank", "count", "year", "name",
}


def is_header_row(row_values):
    """检测第一行是否为表头行（而非标题行）。

    判定规则：
    1. 第一列值命中已知关键词 → 表头
    2. 全行有 ≥2 个值命中关键词 → 表头
    """
    if not row_values:
        return False
    first = str(row_values[0]).strip() if row_values[0] else ""
    if not first:
        return False
    # 快速路径：第一列直接命中
    if first in KNOWN_COLUMN_KEYWORDS:
        return True
    for kw in KNOWN_COLUMN_KEYWORDS:
        if kw in first:
            return True
    # 统计命中数
    hits = 0
    for v in row_values:
        if v is None:
            continue
        sv = str(v).strip()
        if sv in KNOWN_COLUMN_KEYWORDS:
            hits += 1
        else:
            for kw in KNOWN_COLUMN_KEYWORDS:
                if kw in sv:
                    hits += 1
                    break
    return hits >= 2


def read_sheet_dicts(ws, skip_title_row=True):
    """Read worksheet rows as list of dicts.

    Auto-detects and skips a title row (row 1) when the actual header
    is on row 2.  Uses bounds-safe dict construction.

    Args:
        ws: openpyxl worksheet
        skip_title_row: if True, auto-detect and skip title rows before the header row

    Returns:
        List of dicts mapping header → cell value. Empty list if only headers.
    """
    if ws.max_row < 2:
        return []
    header_row_idx = 1
    if skip_title_row and ws.max_row >= 3:
        row1_vals = tuple(cell.value for cell in ws[1])
        if not is_header_row(row1_vals):
            row2_vals = tuple(cell.value for cell in ws[2]) if ws.max_row >= 3 else None
            if row2_vals and is_header_row(row2_vals):
                header_row_idx = 2
    headers = [
        str(cell.value) if cell.value is not None else f"col_{i}"
        for i, cell in enumerate(ws[header_row_idx])
    ]
    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        d = {}
        for i, val in enumerate(row):
            if i < len(headers):
                d[headers[i]] = val
        rows.append(d)
    return rows


# 宏观数据 Sheet 名关键词映射
SHEET_KEY_MAP = {
    "特控线": "特控线",
    "一分一段": "一分一段表",
    "赋分区间": "赋分区间",
    "对照": "本校对照表_总分",
    "门槛": "门槛",
    "升级": "升级",
    "院校层次": "院校层次",
}


def find_sheet(sheets, keyword):
    """Find first sheet name containing keyword (case-insensitive fuzzy match).

    Uses SHEET_KEY_MAP display name for matching when it differs from keyword
    to avoid false matches (e.g. '对照' matching '单科对照' instead of '本校对照表_总分').
    """
    display = SHEET_KEY_MAP.get(keyword, keyword)
    # Prefer display_name match (more specific)
    if display.lower() != keyword.lower():
        for name in sheets:
            if display.lower() in name.lower():
                return name
    # Fallback: sheets matching keyword but excluding known false positives
    for name in sheets:
        nl = name.lower()
        if keyword.lower() in nl and "单科" not in nl:
            return name
    return None


def read_macro_data(workspace):
    """Read all macro data sheets with fuzzy sheet name matching.

    Returns None if the macro Excel file does not exist.
    """
    path = os.path.join(workspace, "data", "macro", "宏观数据_只读.xlsx")
    if not os.path.exists(path):
        path = os.path.join(workspace, "data", "macro", "宏观数据.xlsx")
    if not os.path.exists(path):
        return None
    wb = load_workbook(path, data_only=True)
    data = {}
    # 先用模糊匹配定位关键 Sheet
    matched = set()
    for key, display_name in SHEET_KEY_MAP.items():
        found = find_sheet(wb.sheetnames, key)
        if found:
            data[display_name] = read_sheet_dicts(wb[found])
            matched.add(found)
    # 其余 Sheet 保留原名
    for name in wb.sheetnames:
        if name not in matched:
            data[name] = read_sheet_dicts(wb[name])
    wb.close()
    return data


def filter_numeric_rows(rows, key_field):
    """Filter rows to only those where key_field is numeric (int or float)."""
    return [r for r in rows if isinstance(r.get(key_field), (int, float))]


def filter_score_table(rows):
    """Filter 一分一段表 rows: require both '分数' and '累计人数' to be numeric."""
    return [r for r in filter_numeric_rows(rows, "累计人数") if isinstance(r.get("分数"), (int, float))]
