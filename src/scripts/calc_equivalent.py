#!/usr/bin/env python3
"""Calculate equivalent Gaokao score using all available methods.

Priority order (fixed ranking, first available wins as primary):
  1. 双模块换算法 (Two-module) — A/B级
  2. 人数校准法 (Population calibration) — B级
  3. 分数线对照法/等比例放缩法 (Score-line comparison) — A级
  4. 校排阈值估算法 (School threshold estimation) — B级
  5. 校内排名对照法 (School ranking lookup) — A级
  6. 单科排名对照法 (School subject lookup) — A级
  7. 排名锚定法 (Percentile anchoring) — A级，交叉验证
  8. 校排名估算 (School rank estimation) — C级（低精度回退，不参与融合）

Confidence is A/B/C/D four levels, determined by data source and method.
All A/B-level methods participate in weighted fusion by confidence.
C-level methods appear in cross-validations for transparency but are excluded from fusion.
Weights (A=1.0, B=0.8, C=0.5, D=0).

Input: JSON via stdin
Output: JSON with equivalent score, confidence, error range, cross-validations
"""

import json
import os
import re
import sys

from openpyxl import load_workbook

from config import *  # noqa: F401,F403 — 统一业务常量
from excel_utils import read_macro_data, filter_numeric_rows, filter_score_table


# ── 共享工具函数（抽取自多个方法中的重复逻辑） ────────────────────────


def safe_float(val, default=None):
    """安全转换为 float，失败时返回 default。排除 bool 类型。"""
    if isinstance(val, bool):
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_int(val, default=None):
    """安全转换为 int，失败时返回 default。排除 bool 类型。"""
    if isinstance(val, bool):
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def normalize_confidence(conf):
    """归一化置信度：去除"级"后缀，非 ABCD 值默认为 B。"""
    if not conf:
        return "B"
    conf = str(conf).replace("级", "").strip().upper()
    return conf if conf in ("A", "B", "C", "D") else "B"


def find_latest_gaokao_special_line(special_lines):
    """从特控线数据中查找最新年份的高考特控线。

    消除 3 处重复逻辑：method_score_line / method_population_calibration /
    compute_independent_subject_sum。

    Returns:
        (year, score) 元组，或 (None, None)。
    """
    latest_year = -1
    gaokao_sl = None
    for sl in special_lines:
        try:
            year = int(sl.get("年份", 0))
        except (ValueError, TypeError):
            continue
        sl_val = sl.get("特控线分数")
        if sl_val is None:
            continue
        try:
            sl_f = float(sl_val)
        except (ValueError, TypeError):
            continue
        if year > latest_year:
            latest_year = year
            gaokao_sl = sl_f
    return (gaokao_sl, latest_year if latest_year > 0 else None)


def parse_upgrade_sheet(ws):
    """解析升级 Sheet 的非标准布局（特控线分段 + 浙大线分段）。

    消除 method_two_module / method_school_threshold 中的重复解析逻辑。
    Sheet 结构：标记行（含"特控"/"浙大"+"分段"）后跟数据行
    数据行: col0=科目, col1=2027划线, col2=2027上线, col3=2028划线, col4=2028上线

    Returns:
        dict: {subject: {"special": line_val, "zd": line_val or None,
                          "special_rank": count_val or None, "zd_rank": count_val or None}}
    """
    cutoffs = {}
    current_section = None  # "special" or "zd"

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row:
            continue
        text_0 = str(row[0]).strip() if row[0] else ""

        if "特控" in text_0 and "分段" in text_0:
            current_section = "special"
            continue
        if "浙大" in text_0 and "分段" in text_0:
            current_section = "zd"
            continue
        if current_section is None or len(row) < 5:
            continue

        subj = text_0
        if subj == "科目" or not subj:
            continue

        # col3 = 2028划线 (分数线) — 解析失败时设为 None 而非跳过整行
        line_2028 = None
        if row[3] is not None:
            try:
                line_2028 = float(row[3])
            except (ValueError, TypeError):
                pass

        # col4 = 2028上线 (人数)
        count_2028 = None
        if row[4] is not None:
            try:
                count_2028 = int(row[4])
            except (ValueError, TypeError):
                pass

        if line_2028 is None and count_2028 is None:
            continue

        if subj not in cutoffs:
            cutoffs[subj] = {"special": None, "zd": None, "special_rank": None, "zd_rank": None}

        if line_2028 is not None:
            cutoffs[subj][current_section] = line_2028
        if count_2028 is not None:
            cutoffs[subj][f"{current_section}_rank"] = count_2028

    return cutoffs


def load_upgrade_sheet(workspace, exam_name):
    """加载宏观数据中的升级Sheet（非标准布局，直接读取）。

    按考试名称多级匹配升级Sheet：
      1. 期末考试 → 含"期末"+"升级"的Sheet
      2. 其他考试关键词（期中/月考/联考/模拟/统考）→ 匹配关键词或"升级"
      3. 兜底：任意含"升级"的Sheet

    Returns:
        (wb, ws) 元组；若未找到则返回 (None, None)。
        调用方负责在使用完毕后调用 wb.close()。
    """
    for fname in ("宏观数据_只读.xlsx", "宏观数据.xlsx"):
        path = os.path.join(workspace, "data", "macro", fname)
        if os.path.exists(path):
            break
    else:
        return None, None

    wb = None
    try:
        wb = load_workbook(path, data_only=True)
        ws = None
        # Priority 1: both "期末" in exam and sheet
        if "期末" in exam_name:
            for sn in wb.sheetnames:
                if "期末" in str(sn) and "升级" in str(sn):
                    ws = wb[sn]
                    break
        # Priority 2: match other exam keywords (期中, 月考, 联考, etc.)
        if ws is None:
            for kw in EXAM_KEYWORDS:
                if kw == "期末":
                    continue
                if kw in exam_name:
                    for sn in wb.sheetnames:
                        if kw in str(sn) or "升级" in str(sn):
                            ws = wb[sn]
                            break
                    if ws:
                        break
        # Priority 3: any sheet with "升级" as fallback
        if ws is None:
            for sn in wb.sheetnames:
                if "升级" in str(sn):
                    ws = wb[sn]
                    break
        if ws is None:
            wb.close()
            return None, None
        return wb, ws
    except Exception:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        return None, None


def load_upgrade_cutoffs(data):
    """加载并解析升级Sheet，返回 cutoffs dict 或 None。

    消除 method_two_module / method_school_threshold 中的重复加载逻辑。
    负责 workbook 的完整生命周期管理。
    """
    workspace = os.path.abspath(data.get("workspace", "."))
    exam_name = data.get("exam_name", "")
    wb, ws = load_upgrade_sheet(workspace, exam_name)
    if ws is None:
        return None
    try:
        return parse_upgrade_sheet(ws)
    except Exception:
        return None
    finally:
        wb.close()


def compute_main_raw_sum(data):
    """计算语数英原始分总和；若不足三科则按比例折算。

    优先使用 subjects 中的语数英原始分；若三科不全，则按总分比例
    折算到语数英满分制。

    Returns:
        语数英等效原始分总和（float）。
    """
    subjects_input = data.get("subjects", [])
    main_raw_sum = 0.0
    main_count = 0
    for subj in subjects_input:
        name = subj.get("name", "")
        raw = subj.get("raw")
        if name in MAIN_SUBJECTS and raw is not None and raw != "":
            try:
                main_raw_sum += float(raw)
                main_count += 1
            except (ValueError, TypeError):
                continue

    if main_count == 3:  # All three 语数英 scores present — use actual
        return main_raw_sum
    else:
        student_score = safe_float(data.get("_original_total_score", data.get("total_score", 0)), 0)
        if student_score is None:
            student_score = 0.0
        score_scale = safe_int(data.get("score_scale", FULL_SCORE), FULL_SCORE)
        if score_scale is None:
            score_scale = FULL_SCORE
        if score_scale == FULL_SCORE:
            return student_score * MAIN_FULL_SCORE / FULL_SCORE
        else:
            return student_score


def find_score_by_count(sorted_table, target_count):
    """在一分一段表中查找累计人数最接近 target_count 的分数。

    Args:
        sorted_table: 已排序的一分一段表（list of dict，含"分数"和"累计人数"）。
        target_count: 目标累计人数。

    Returns:
        最接近的分数（float），或 None（表为空时）。
    """
    best_score = None
    best_diff = float("inf")
    for row in sorted_table:
        try:
            count = int(row.get("累计人数", 0))
            score = float(row["分数"])
        except (ValueError, TypeError, KeyError):
            continue
        diff = abs(count - target_count)
        if diff < best_diff:
            best_diff = diff
            best_score = score
    return best_score


def method_score_line(data, macro):
    """优先级3: 分数线对照法（等比例放缩）— A级."""
    special_line_exam = data.get("special_line_exam") or data.get("special_line")
    if not special_line_exam:
        return None

    special_line_exam = safe_float(special_line_exam)
    if special_line_exam is None:
        return None

    special_lines = filter_numeric_rows(macro.get("特控线", []), "特控线分数")
    if not special_lines:
        return None

    # Find the most recent gaokao special line (by year)
    gaokao_sl, _ = find_latest_gaokao_special_line(special_lines)

    if not gaokao_sl:
        return None

    total_score = safe_float(data.get("total_score"))
    if total_score is None or total_score == 0:
        return None

    if total_score == FULL_SCORE:
        return {
            "method": "分数线对照法",
            "score": float(FULL_SCORE),
            "confidence": "A",
            "detail": f"满分{FULL_SCORE} → 等效分{FULL_SCORE}",
        }

    if special_line_exam >= FULL_SCORE:
        return None

    es = (FULL_SCORE - gaokao_sl) / (FULL_SCORE - special_line_exam) * (total_score - special_line_exam) + gaokao_sl
    return {
        "method": "分数线对照法",
        "score": round(es, 1),
        "confidence": "A",
        "detail": f"等效分 = ({FULL_SCORE}-{gaokao_sl})/({FULL_SCORE}-{special_line_exam})×({total_score}-{special_line_exam})+{gaokao_sl} = {es:.1f}",
    }


def method_school_lookup(data, macro):
    """优先级5: 校内排名对照法 — A级."""
    school_rank = data.get("school_rank")
    if not school_rank:
        return None

    unexamined_top = safe_int(data.get("unexamined_top_students", 0), 0) or 0

    # Apply class-type calibration
    sr = safe_int(school_rank)
    if sr is None:
        return None
    calibrated_rank = sr + unexamined_top

    lookup_sheet = macro.get("本校对照表_总分", [])
    if not lookup_sheet:
        return None

    # Validate required columns exist in first row
    if "校内排名" not in lookup_sheet[0]:
        return None

    # Filter to rows with numeric 校内排名 and 高考总分, then sort
    lookup = []
    for r in lookup_sheet:
        try:
            int(r.get("校内排名"))
            float(r.get("高考总分"))
            lookup.append(r)
        except (ValueError, TypeError):
            continue
    if not lookup:
        return None
    lookup = sorted(lookup, key=lambda r: int(r.get("校内排名", 0)))

    ranks = [int(r["校内排名"]) for r in lookup]
    scores = [float(r["高考总分"]) for r in lookup]

    # Exact match
    if calibrated_rank in ranks:
        idx = ranks.index(calibrated_rank)
        score = scores[idx]
    elif calibrated_rank < ranks[0]:
        score = scores[0]
    elif calibrated_rank > ranks[-1]:
        score = scores[-1]
    else:
        # Linear interpolation
        for i in range(len(ranks) - 1):
            if ranks[i] <= calibrated_rank <= ranks[i + 1]:
                denom = ranks[i + 1] - ranks[i]
                if denom == 0:
                    score = scores[i]
                else:
                    ratio = (calibrated_rank - ranks[i]) / denom
                    score = scores[i] + ratio * (scores[i + 1] - scores[i])
                break
        else:
            return None

    detail = f"校内排名{calibrated_rank}名"
    if unexamined_top:
        detail += f"（原始排名{school_rank}，补算重点班{unexamined_top}人）"
    detail += f" → 对照表对应高考总分{score:.0f}分"

    return {
        "method": "校内排名对照法",
        "score": round(score, 1),
        "confidence": "A",
        "detail": detail,
    }


def method_percentile(data, macro):
    """优先级7: 百分位排名锚定法 — A级，交叉验证."""
    rank = data.get("city_rank") or data.get("alliance_rank")
    total = data.get("city_total") or data.get("alliance_total")

    if not rank or not total:
        return None

    rank = safe_int(rank)
    total = safe_int(total)
    if rank is None or total is None:
        return None

    if rank <= 0 or total <= 0 or rank > total:
        return None

    percentile = 1.0 - (rank / total)
    score_table = filter_score_table(macro.get("一分一段表", []))
    if not score_table:
        return None

    # Sort by 分数 descending (higher score = lower cumulative count)
    sorted_table = sorted(score_table, key=lambda r: int(r.get("分数", 0)), reverse=True)

    max_count = max(int(r.get("累计人数", 0)) for r in sorted_table)
    if max_count == 0:
        return None

    target_count = int((1 - percentile) * max_count)

    # Find closest match
    best_score = find_score_by_count(sorted_table, target_count)

    if best_score is None:
        return None

    source = "全市排名" if data.get("city_rank") else "联盟排名"
    return {
        "method": "排名锚定法",
        "score": round(best_score, 1),
        "confidence": "A",
        "detail": f"{source}{rank}/{total} → 百分位{percentile:.3f} → 等效分{best_score:.0f}",
    }


def method_school_estimate(data, macro):
    """优先级8: 校排名估算 — C级（低精度回退）.

    仅校内排名、无本校对照表时使用。用学校类型系数估算全市排名，
    再通过一分一段表百分位锚定得到等效分。

    注意：此方法置信度为C级，被置信度门槛拦截（仅C级时返回insufficient_data），
    但在有更高级方法时仍参与交叉验证以提供透明度。不参与加权融合。
    """
    school_rank = data.get("school_rank")
    school_total = data.get("school_total")
    if not school_rank or not school_total:
        return None

    sr = safe_int(school_rank)
    st = safe_int(school_total)
    if sr is None or st is None or st == 0:
        return None

    # 有对照表时应走 method_school_lookup
    if macro.get("本校对照表_总分"):
        return None

    score_table = filter_score_table(macro.get("一分一段表", []))
    if not score_table:
        return None
    max_count = max(int(r.get("累计人数", 0)) for r in score_table)
    if max_count == 0:
        return None

    # 学校类型系数
    school_type = data.get("school_type", "普通")
    coeff = SCHOOL_TYPE_COEFF.get(school_type, 1.0)

    estimated_city_rank = int(sr / st * max_count * coeff)
    estimated_city_rank = min(estimated_city_rank, max_count)  # clamp to avoid percentile overflow
    percentile = 1.0 - (estimated_city_rank / max_count)
    percentile = max(0.0, min(1.0, percentile))  # clamp to [0, 1]

    sorted_table = sorted(score_table, key=lambda r: int(r.get("分数", 0)), reverse=True)
    target_count = int((1 - percentile) * max_count)
    best_score = find_score_by_count(sorted_table, target_count)

    if best_score is None:
        return None

    return {
        "method": "校排名估算",
        "score": round(best_score, 1),
        "confidence": "C",
        "detail": f"校内排名{school_rank}/{school_total}（{school_type or '未知类型'}）→ 估算全市排名~{estimated_city_rank} → 等效分{best_score:.0f}",
    }


def method_population_calibration(data, macro):
    """优先级2: 人数校准法 (Population calibration) — B级.

    利用校内门槛上线人数与高考一分一段表的人数映射关系，
    将校内排名转换为高考排名，再查一分一段表得到等效分。

    适用条件：有校内排名 + 门槛Sheet中有"特控线上线人数"
    置信度：B级（高于C级校排名估算，低于A级分数线对照）
    """
    school_rank = data.get("school_rank")
    if not school_rank:
        return None

    # 应用班型校准（重点班未参考人数），与 method_school_lookup 一致
    unexamined_top = safe_int(data.get("unexamined_top_students", 0), 0) or 0
    sr = safe_int(school_rank)
    if sr is None:
        return None
    calibrated_rank = sr + unexamined_top

    # 从门槛Sheet读取特控线上线人数
    threshold_sheet = macro.get("门槛") or macro.get("升级")
    if not threshold_sheet:
        # 兼容直接传入的 macro dict（测试或非标准键名）
        for key in macro:
            if "门槛" in str(key) or "升级" in str(key):
                threshold_sheet = macro[key]
                break
    if not threshold_sheet:
        return None

    school_teckong_count = None
    for row in threshold_sheet:
        for k, v in row.items():
            if "特控" in str(k) and "上线" in str(k):
                count = safe_int(v)
                if count is not None:
                    school_teckong_count = count
                    break
        if school_teckong_count:
            break

    if not school_teckong_count:
        return None

    # 高考特控线对应人数（从一分一段表查）
    special_lines = filter_numeric_rows(macro.get("特控线", []), "特控线分数")
    if not special_lines:
        return None

    gaokao_line, _ = find_latest_gaokao_special_line(special_lines)

    if not gaokao_line:
        return None

    score_table = filter_score_table(macro.get("一分一段表", []))
    if not score_table:
        return None

    # 查找高考特控线对应的累计人数
    sorted_table = sorted(score_table, key=lambda r: int(r.get("分数", 0)), reverse=True)
    gaokao_count = None
    best_diff = float("inf")
    for row in sorted_table:
        score = float(row.get("分数", 0))
        diff = abs(score - gaokao_line)
        if diff < best_diff:
            best_diff = diff
            gaokao_count = int(row.get("累计人数", 0))

    if not gaokao_count:
        return None

    # 校准系数
    k = gaokao_count / school_teckong_count
    city_rank = int(calibrated_rank * k)
    max_count = max(int(r.get("累计人数", 0)) for r in sorted_table)
    city_rank = min(city_rank, max_count)

    # 在排序表中查找最接近的排名对应的分数
    best_score = find_score_by_count(sorted_table, city_rank)

    if best_score is None:
        return None

    return {
        "method": "人数校准法",
        "score": round(best_score, 1),
        "confidence": "B",
        "detail": f"校内排名{school_rank}（含重点班校准{unexamined_top}人） × k({k:.1f}) → 高考排名{city_rank} → 等效分{best_score:.0f}",
    }


def method_two_module(data, macro):
    """优先级1: 双模块独立换算法 — A/B级，最高优先级.

    When school upgrade data (per-subject 特控线+浙大线) is available,
    splits calculation into two independent modules:

    Module 1 (语数英 450pt): proportional scaling between school cutoffs
      and gaokao reference targets (特控341, 浙大382).

    Module 2 (选科 100pt each): per-subject proportional scaling.
      Priority 1: dual-line (特控+浙大) → A级
      Priority 2: single-line (仅特控) → B级
      Priority 3: no line → skip (delegated to fallback methods)

    Total = 语数英等效 + sum(选科等效)
    """
    # Load upgrade sheet using shared helper
    cutoffs = load_upgrade_cutoffs(data)
    if cutoffs is None:
        return None

    # Need at least 语数英综合 special line
    main_data = cutoffs.get("语数英综合", {})
    if main_data.get("special") is None:
        return None

    # Gaokao reference targets (from config)
    GK_MAIN_SPECIAL = GAOKAO_TARGETS["main_special"]
    GK_MAIN_ZD = GAOKAO_TARGETS["main_zd"]
    GK_SUB_SPECIAL = GAOKAO_TARGETS["sub_special"]
    GK_SUB_ZD = GAOKAO_TARGETS["sub_zd"]

    # Calculate actual 语数英 raw score sum from subjects; fall back to proportional
    student_main = compute_main_raw_sum(data)

    details = []
    total_equivalent = 0.0
    conf_counts = {"A": 0, "B": 0, "C": 0, "D": 0}  # per-module confidence tally

    # ── Module 1: 语数英 ──
    sch_special = main_data["special"]
    sch_zd = main_data.get("zd")

    if sch_zd is not None and sch_zd != sch_special and student_main >= sch_special:
        # Clamp ratio to [0, 1] for linear interpolation between special and ZJU
        denom = sch_zd - sch_special
        ratio = (student_main - sch_special) / denom
        ratio_clamped = min(max(ratio, 0.0), 1.0)
        main_eq = GK_MAIN_SPECIAL + (GK_MAIN_ZD - GK_MAIN_SPECIAL) * ratio_clamped
        conf = "A"

        # Diminishing returns for excess above ZJU line
        if student_main > sch_zd:
            excess = student_main - sch_zd
            per_point = (GK_MAIN_ZD - GK_MAIN_SPECIAL) / denom
            bonus = excess * DAMPING * per_point
            main_eq = min(main_eq + bonus, MAIN_MAX)
            detail = (f"语数英{student_main:.0f}分, 校特控{sch_special:.0f}/浙大{sch_zd:.0f}"
                      f" → 线上{ratio_clamped:.0%}(封顶)+超额衰减×{DAMPING} → 等效{main_eq:.1f}")
        else:
            detail = (f"语数英{student_main:.0f}分, 校特控{sch_special:.0f}/浙大{sch_zd:.0f}"
                      f" → 线上{ratio:.1%} → 等效{main_eq:.1f}")
    elif student_main >= sch_special and sch_special > 0:
        # Only special line available — single-point scaling
        ratio = student_main / sch_special
        ratio_clamped = min(ratio, 1.0)
        main_eq = GK_MAIN_SPECIAL * ratio_clamped
        conf = "B"

        # Diminishing returns for excess above special line
        if student_main > sch_special:
            excess = student_main - sch_special
            per_point = GK_MAIN_SPECIAL / sch_special
            bonus = excess * DAMPING * per_point
            main_eq = min(main_eq + bonus, MAIN_MAX)
            detail = (f"语数英{student_main:.0f}分, 校特控{sch_special:.0f}(无浙大线)"
                      f" → 比例{ratio_clamped:.0%}(封顶)+超额衰减×{DAMPING} → 等效{main_eq:.1f}")
        else:
            detail = (f"语数英{student_main:.0f}分, 校特控{sch_special:.0f}(无浙大线)"
                      f" → 比例{ratio:.1%} → 等效{main_eq:.1f}")
    else:
        return None  # Below special line, can't use this method

    main_eq = round(main_eq, 1)
    total_equivalent += main_eq
    conf_counts[conf] += 1
    details.append(f"[语数英] {detail}")

    # ── Module 2: 选科 ──
    workspace = os.path.abspath(data.get("workspace", "."))
    subjects_input = data.get("subjects", [])
    for subj in subjects_input:
        name = subj.get("name", "")
        if name in MAIN_SUBJECTS:
            continue  # handled in module 1
        raw = subj.get("raw")
        if raw is None or raw == "":
            details.append(f"[{name}] 无原始分, 跳过")
            continue

        raw = safe_float(raw)
        if raw is None:
            details.append(f"[{name}] 原始分格式异常, 跳过")
            continue
        sub_cut = cutoffs.get(name, {})

        if sub_cut.get("special") and sub_cut.get("zd") and sub_cut["zd"] != sub_cut["special"] and raw >= sub_cut["special"] and sub_cut["special"] > 0:
            # Priority 1: dual-line — clamp ratio to [0,1], diminishing returns above ZJU
            sub_denom = sub_cut["zd"] - sub_cut["special"]
            ratio = (raw - sub_cut["special"]) / sub_denom
            ratio_clamped = min(max(ratio, 0.0), 1.0)
            sub_eq = GK_SUB_SPECIAL + (GK_SUB_ZD - GK_SUB_SPECIAL) * ratio_clamped
            conf = "A"
            if raw > sub_cut["zd"]:
                excess = raw - sub_cut["zd"]
                per_point = (GK_SUB_ZD - GK_SUB_SPECIAL) / sub_denom
                bonus = excess * DAMPING * per_point
                sub_eq = min(sub_eq + bonus, SUB_MAX)
                detail = (f"{name}{raw:.0f}分, 校特控{sub_cut['special']:.0f}/浙大{sub_cut['zd']:.0f}"
                          f" → 线上{ratio_clamped:.0%}(封顶)+超额衰减×{DAMPING} → 等效{sub_eq:.1f}")
            else:
                detail = (f"{name}{raw:.0f}分, 校特控{sub_cut['special']:.0f}/浙大{sub_cut['zd']:.0f}"
                          f" → 线上{ratio:.1%} → 等效{sub_eq:.1f}")
        elif sub_cut.get("special") and raw >= sub_cut["special"] and sub_cut["special"] > 0:
            # Priority 2: single-line — clamp ratio, diminishing returns above special
            ratio = raw / sub_cut["special"]
            ratio_clamped = min(ratio, 1.0)
            sub_eq = GK_SUB_SPECIAL * ratio_clamped
            conf = "B"
            if raw > sub_cut["special"]:
                excess = raw - sub_cut["special"]
                per_point = GK_SUB_SPECIAL / sub_cut["special"]
                bonus = excess * DAMPING * per_point
                sub_eq = min(sub_eq + bonus, SUB_MAX)
                detail = (f"{name}{raw:.0f}分, 校特控{sub_cut['special']:.0f}(无浙大线)"
                          f" → 比例{ratio_clamped:.0%}(封顶)+超额衰减×{DAMPING} → 等效{sub_eq:.1f}")
            else:
                detail = (f"{name}{raw:.0f}分, 校特控{sub_cut['special']:.0f}(无浙大线)"
                          f" → 比例{ratio:.1%} → 等效{sub_eq:.1f}")
        elif sub_cut.get("special"):
            # Below special line
            ratio = raw / sub_cut["special"]
            sub_eq = GK_SUB_SPECIAL * ratio
            conf = "C"
            detail = (f"{name}{raw:.0f}分, 低于校特控{sub_cut['special']:.0f}"
                      f" → 比例{ratio:.1%} → 等效{sub_eq:.1f}")
        else:
            # Priority 3: no school cutoff → try existing single-subject fallbacks
            assigned = subj.get("assigned")
            if assigned is not None and assigned != "":
                sub_eq = safe_float(assigned)
                if sub_eq is None:
                    sub_eq = 0.0
                conf = normalize_confidence(subj.get("confidence", "B"))
                detail = f"{name}赋分{assigned}（{conf}级）→ 等效{sub_eq:.0f}分"
            else:
                # Try cross-exam fallback
                prev = _find_previous_subject_data(workspace, name, data.get("exam_name", ""))
                if prev:
                    n = prev.get("exams_skipped", 1)
                    discount = round(CROSS_EXAM_DISCOUNT ** n, 3)
                    sub_eq = round(prev["score"] * discount, 1)
                    conf = "C"
                    detail = (f"{name}无校内划线, 回退至{prev['exam']}"
                              f"（{prev['score']}分×{discount:.2f}={sub_eq}分, C级）")
                else:
                    # Use school对照表 or rough estimate
                    sub_eq = round(GK_SUB_SPECIAL * raw / SUB_FULL_SCORE, 1)
                    conf = "D"
                    detail = f"{name}无任何参照数据 → 粗略估算{sub_eq}分(D级)"
            if sub_eq is None:
                details.append(f"[{name}] 无任何可用数据, 跳过")
                continue

        sub_eq = round(sub_eq, 1)
        total_equivalent += sub_eq
        conf_counts[conf] += 1
        details.append(f"[{name}] {detail}")

    total_equivalent = round(total_equivalent, 1)

    # Overall confidence: A if ≥50% modules are A and no module is D
    total_modules = sum(conf_counts.values())
    if conf_counts.get("D", 0) > 0:
        overall_conf = "C"
    elif conf_counts.get("C", 0) >= 2:
        overall_conf = "B"
    elif conf_counts.get("A", 0) >= total_modules * 0.5:
        overall_conf = "A"
    elif conf_counts.get("A", 0) + conf_counts.get("B", 0) >= total_modules * 0.5:
        overall_conf = "B"
    else:
        overall_conf = "C"

    return {
        "method": "双模块换算法",
        "score": total_equivalent,
        "confidence": overall_conf,
        "detail": " | ".join(details),
    }


def method_school_threshold(data, macro):
    """优先级4: 校排阈值估算法 — B级，交叉验证.

    Reads the升级 Sheet directly (bypasses broken dict-parsing for
    non-standard sheet layout). Uses school-internal 特控线+浙大线
    thresholds to estimate school rank → 一分一段表 → equivalent.
    """
    # Load upgrade sheet using shared helper
    cutoffs = load_upgrade_cutoffs(data)
    if cutoffs is None:
        return None

    # Extract 语数英综合 特控线/浙大线 及对应上线人数
    main_data = cutoffs.get("语数英综合", {})
    te_line = main_data.get("special")
    zd_line = main_data.get("zd")
    te_rank = main_data.get("special_rank")
    zd_rank = main_data.get("zd_rank")

    if te_line is None or zd_line is None or te_rank is None or zd_rank is None:
        return None

    # Use actual 语数英 raw score sum from subjects; fall back to proportional
    student_450 = compute_main_raw_sum(data)

    if te_line >= zd_line:
        return None  # 特控线应低于浙大线，数据异常

    if not (te_line <= student_450 <= zd_line):
        return None

    denom = zd_line - te_line
    if denom == 0:
        return None  # 除零保护
    ratio = (student_450 - te_line) / denom
    estimated_rank = int(te_rank - ratio * (te_rank - zd_rank))

    # Get school total: prefer input data, then try macro sheets
    school_total = data.get("school_total")
    if not school_total:
        # Try to read from macro data (any sheet with "结构" in name)
        for key in macro:
            if "结构" in str(key):
                for row in macro[key]:
                    vals = list(row.values())
                    if len(vals) >= 2:
                        text = str(vals[0]) if vals[0] else ""
                        if "全校" in text:
                            try:
                                school_total = int(str(vals[1]).replace("人", ""))
                            except (ValueError, TypeError):
                                pass
                if school_total:
                    break  # 仅在找到有效数据后退出外层循环
    st = safe_int(school_total)
    if st is None or st == 0:
        return None  # 无法估算，缺少学校总人数

    school_type = data.get("school_type", "普通")
    coeff = SCHOOL_TYPE_COEFF.get(school_type, 1.0)

    score_table = filter_score_table(macro.get("一分一段表", []))
    if not score_table:
        return None
    max_count = max(int(r.get("累计人数", 0)) for r in score_table)
    if max_count == 0:
        return None

    estimated_city_rank = int(estimated_rank / st * max_count * coeff)
    estimated_city_rank = min(estimated_city_rank, max_count)
    percentile = max(0.0, min(1.0, 1.0 - estimated_city_rank / max_count))

    sorted_table = sorted(score_table, key=lambda r: int(r.get("分数", 0)), reverse=True)
    target_count = int((1 - percentile) * max_count)
    best_score = find_score_by_count(sorted_table, target_count)

    if best_score is None:
        return None

    return {
        "method": "校排阈值估算法",
        "score": round(best_score, 1),
        "confidence": "B",
        "detail": f"校内特控线{te_line:.0f}分(=第{te_rank}名), 浙大线{zd_line:.0f}分(=第{zd_rank}名) → 学生{student_450:.0f}分估算校内第{estimated_rank}名 → 等效{best_score:.0f}分",
    }


def method_school_subject_lookup(data, macro):
    """优先级6: 单科排名对照法 — A级.

    利用宏观数据中的"单科对照"Sheet（校内单科排名→高考等效分映射表），
    对每个选科用校内单科排名查表得到等效分。
    """
    subject_rank_data = read_school_subject_data(macro)
    if not subject_rank_data:
        return None

    subjects_input = data.get("subjects", [])
    subject_results = []
    for subj in subjects_input:
        name = subj.get("name", "")
        subj_rank = subj.get("school_rank")
        if not subj_rank or name not in subject_rank_data:
            continue
        rank_map = subject_rank_data[name].get("rank_scores", {})
        if not rank_map:
            continue
        subj_rank = int(subj_rank)
        # 线性插值查找
        ranks = sorted(rank_map.keys())
        if not ranks:
            continue
        if subj_rank <= ranks[0]:
            eq_score = round(rank_map[ranks[0]], 1)
        elif subj_rank >= ranks[-1]:
            eq_score = round(rank_map[ranks[-1]], 1)
        else:
            for i in range(len(ranks) - 1):
                if ranks[i] <= subj_rank <= ranks[i + 1]:
                    denom = ranks[i + 1] - ranks[i]
                    if denom == 0:
                        eq_score = round(rank_map[ranks[i]], 1)
                    else:
                        ratio = (subj_rank - ranks[i]) / denom
                        eq_score = rank_map[ranks[i]] + ratio * (rank_map[ranks[i + 1]] - rank_map[ranks[i]])
                        eq_score = round(eq_score, 1)
                    break
            else:
                continue
        subject_results.append({
            "name": name,
            "equivalent_score": eq_score,
            "confidence": "A",
            "detail": f"{name}校内排名{subj_rank} → 单科对照表 → 等效分{eq_score:.0f}",
        })

    if not subject_results:
        return None

    # 计算总分等效分 = 各科等效分之和
    total_eq = sum(r["equivalent_score"] for r in subject_results)
    return {
        "method": "单科排名对照法",
        "score": round(total_eq, 1),
        "confidence": "A",
        "detail": " | ".join(r["detail"] for r in subject_results),
        "subject_scores": subject_results,
    }


def read_school_subject_data(macro):
    """Extract per-subject rank→score mappings from 本校对照表 sheets.

    Only processes sheets whose names contain 本+对照+总分 (excluding
    the standard 本校对照表_总分 which is handled by method_school_lookup).
    These are subject-level lookup tables mapping school ranks to gaokao scores.
    Returns {subject: {rank_scores: {rank: score, ...}}}.
    """
    result = {}

    for sheet_key in macro:
        sname = str(sheet_key)
        # 排除标准总分对照表（由 method_school_lookup 使用）
        if sname == "本校对照表_总分":
            continue
        # 匹配单科对照表（含"单科"+"对照"关键字）
        if not ("单科" in sname and "对照" in sname):
            continue

        for row in macro[sheet_key]:
            subj = str(list(row.values())[0]).strip() if row else ""
            if subj not in ALL_SUBJECTS:
                continue
            result.setdefault(subj, {})
            rank_scores = {}
            keys = list(row.keys())
            for k in keys[2:]:  # skip 学科, 参考人数
                val = row.get(k)
                if val and str(val).replace(".", "").replace("-", "").isdigit():
                    try:
                        rank = int(str(k))
                    except (ValueError, TypeError):
                        m = re.search(r'\d+', str(k))
                        if m:
                            rank = int(m.group())
                        else:
                            continue
                    rank_scores[rank] = float(val)
            if rank_scores:
                result[subj]["rank_scores"] = rank_scores

    return result


def compute_subject_equivalents(data, macro):
    """Compute per-subject equivalent scores.

    Two-pass approach:
    1. Resolve ALL 选科 scores first (赋分直映 → fallback → school lookup).
    2. Compute 语数英 from the remaining equivalent (total − sum of 选科).

    This ensures 语数英 + 选科 ≈ total_equivalent regardless of score scale.
    """
    subjects_input = data.get("subjects", [])
    if not subjects_input:
        return []

    total_equivalent = data.get("_total_equivalent", 0)
    workspace = os.path.abspath(data.get("workspace", "."))
    exam_name = data.get("exam_name", "")

    school_data = read_school_subject_data(macro)
    results = []

    # ── Pass 1: resolve 选科 scores ──
    sum_assigned = 0.0
    sum_main_raw = 0.0

    for subj in subjects_input:
        name = subj.get("name", "")
        raw = subj.get("raw")
        assigned = subj.get("assigned")
        confidence = normalize_confidence(subj.get("confidence", "B"))

        if name in MAIN_SUBJECTS:
            if raw is not None and raw != "":
                try:
                    sum_main_raw += float(raw)
                except (ValueError, TypeError):
                    pass
            continue

        # 赋分直映
        if assigned is not None and assigned != "":
            score = safe_float(assigned)
            if score is None:
                continue
            results.append({
                "subject": name, "score": score, "confidence": confidence,
                "method": "赋分直映法",
                "detail": f"{name}赋分{assigned}（{confidence}级）→ 等效高考{assigned}分",
            })
            sum_assigned += score
            continue

        # 无赋分：尝试跨次回退
        prev = _find_previous_subject_data(workspace, name, exam_name)
        if prev:
            n = prev.get("exams_skipped", 1)
            discount = round(CROSS_EXAM_DISCOUNT ** n, 3)
            score = round(prev["score"] * discount, 1)
            discount_pct = f"{discount:.2f}"
            results.append({
                "subject": name, "score": score, "confidence": "C",
                "method": f"数据不足（回退至{prev['exam']}）",
                "detail": f"本次考试缺少{name}数据，回退至{prev['exam']}的数据（{prev['score']}分×{discount_pct}={score}分，C级）。建议补录后重新计算。",
            })
            sum_assigned += score
            continue

        # 无回退：尝试校内对照
        if name in school_data:
            sd = school_data[name]
            if "rank_scores" in sd and sd["rank_scores"]:
                avg_gaokao = sum(sd["rank_scores"].values()) / len(sd["rank_scores"])
                score = round(avg_gaokao, 1)
                results.append({
                    "subject": name, "score": score, "confidence": "C",
                    "method": "校内均值参照法",
                    "detail": f"{name}无赋分无排名→ 参照本校历届{name}均分约{avg_gaokao:.0f}分（C级，仅参考）",
                })
                sum_assigned += score
                continue

        # 完全无数据
        results.append({
            "subject": name, "score": None,
            "confidence": confidence if confidence in ("C", "D") else "D",
            "method": "数据不足",
            "detail": f"{name}无赋分无排名数据，无法计算等效分",
        })

    # ── Pass 2: 语数英 — 从剩余等效分中按比例分配 ──
    remaining = max(0, total_equivalent - sum_assigned)
    main_subjects_found = []
    for subj in subjects_input:
        name = subj.get("name", "")
        raw = subj.get("raw")
        if name not in MAIN_SUBJECTS:
            continue
        raw_f = safe_float(raw) if raw is not None and raw != "" else None
        main_subjects_found.append((name, raw_f))

    # 计算有效的语数英原始分总和
    valid_mains = [(name, rf) for name, rf in main_subjects_found if rf is not None and rf > 0]

    if valid_mains and sum(rf for _, rf in valid_mains) > 0:
        sum_valid = sum(rf for _, rf in valid_mains)
        for name, raw_f in main_subjects_found:
            if raw_f is not None and raw_f > 0:
                ratio = raw_f / sum_valid
                eq = round(remaining * ratio, 1)
                results.append({
                    "subject": name, "score": eq, "confidence": "B",
                    "method": "比例折算法",
                    "detail": f"总等效{total_equivalent} - 选科贡献{sum_assigned} = {remaining:.1f}(剩余) -> {name}占语数英{ratio:.1%} -> 等效{eq}分",
                })
            else:
                # 语数英中有原始分缺失的科目：均分剩余额度
                eq = round(remaining / len(main_subjects_found), 1) if main_subjects_found else 0
                results.append({
                    "subject": name, "score": eq, "confidence": "C",
                    "method": "均分估算法",
                    "detail": f"{name}无原始分，按均分估算等效{eq}分(C级)",
                })
    elif main_subjects_found:
        # 所有语数英都无原始分：均分剩余
        eq = round(remaining / len(main_subjects_found), 1)
        for name, _ in main_subjects_found:
            results.append({
                "subject": name, "score": eq, "confidence": "C",
                "method": "均分估算法",
                "detail": f"{name}无原始分，按均分估算等效{eq}分(C级)",
            })

    return results


def _find_previous_subject_data(workspace, subject_name, current_exam_name):
    """Fallback: find the most recent exam with valid data for a subject.

    Reads 成绩总表.xlsx and looks for previous exams that have
    usable data for the given subject.
    Returns dict with exam, score, field, exams_skipped or None.
    exams_skipped counts how many exam rows were skipped before finding data.
    """
    path = os.path.join(workspace, "data", "personal", "成绩总表.xlsx")
    if not os.path.exists(path):
        return None

    wb = None
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb["成绩总表"]
        headers = [c.value for c in ws[1]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        rows.reverse()  # most recent first

        exams_skipped = 0
        for i, row in enumerate(rows):
            d = dict(zip(headers, row))
            exam_name = str(d.get("考试名", ""))
            # Skip current exam: either first row or matching name
            if i == 0:
                exams_skipped = 0
                continue
            if current_exam_name and (current_exam_name in exam_name or exam_name in current_exam_name):
                exams_skipped = 0
                continue

            # Check if this exam has data for this subject
            if subject_name in MAIN_SUBJECTS:
                score = d.get(subject_name)
                if score is not None and score != "":
                    score_f = safe_float(score)
                    if score_f is not None:
                        return {"exam": exam_name, "score": score_f, "field": subject_name, "exams_skipped": exams_skipped}
            else:
                for j in range(1, 4):
                    subj_name = str(d.get(f"选科{j}名称", ""))
                    if subj_name == subject_name:
                        assigned = d.get(f"选科{j}赋分")
                        if assigned is not None and assigned != "":  # 优先赋分
                            assigned_f = safe_float(assigned)
                            if assigned_f is not None:
                                return {"exam": exam_name, "score": assigned_f, "field": f"选科{j}赋分", "exams_skipped": exams_skipped}
                        raw = d.get(f"选科{j}原始分")
                        if raw is not None and raw != "":
                            raw_f = safe_float(raw)
                            if raw_f is not None:
                                return {"exam": exam_name, "score": raw_f, "field": f"选科{j}原始分", "exams_skipped": exams_skipped}
            exams_skipped += 1
    except (KeyError, IOError):
        return None
    except Exception:
        return None
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
    return None


def compute_independent_subject_sum(data, macro):
    """Compute subject sum independently for fusion.

    Uses 分数线对照法 for 语数英 (independent of primary total method)
    and 赋分直映 for 选科. This produces an estimate that can diverge
    from the total methods, making weighted fusion meaningful.

    Returns dict {"sum": float, "confidences": [str]} or None.
    """
    subjects = data.get("subjects", [])
    total_score = safe_float(data.get("total_score", 0), 0)
    if total_score is None:
        total_score = 0.0
    special_line_exam = data.get("special_line_exam") or data.get("special_line")

    if not subjects or not total_score:
        return None

    # 语数英: use 分数线对照法 (same formula as method_score_line)
    main_raw = 0.0
    for subj in subjects:
        name = subj.get("name", "")
        raw = subj.get("raw")
        if name in MAIN_SUBJECTS and raw is not None and raw != "":
            try:
                main_raw += float(raw)
            except (ValueError, TypeError):
                pass

    if main_raw <= 0:
        return None

    special_lines = filter_numeric_rows(macro.get("特控线", []), "特控线分数")
    gaokao_sl, _ = find_latest_gaokao_special_line(special_lines)

    if not gaokao_sl or not special_line_exam:
        return None

    sl_exam = safe_float(special_line_exam)
    if sl_exam is None or sl_exam >= FULL_SCORE:
        return None

    # 分数线对照法 applied to total, then allocate 语数英 portion
    if total_score >= FULL_SCORE:
        total_via_sl = float(FULL_SCORE)
    else:
        total_via_sl = (FULL_SCORE - gaokao_sl) / (FULL_SCORE - sl_exam) * (total_score - sl_exam) + gaokao_sl

    original_total = safe_float(data.get("_original_total_score", total_score), total_score)
    if original_total is None:
        original_total = total_score
    main_ratio = main_raw / original_total if original_total else 0
    main_eq = main_ratio * total_via_sl

    # 选科: 赋分直映
    subject_sum = main_eq
    confidences = []
    # 语数英 via 分数线对照法 → A级 (3 subjects, only if raw is valid)
    for subj in subjects:
        name = subj.get("name", "")
        raw = subj.get("raw")
        if name in MAIN_SUBJECTS and raw is not None and raw != "":
            confidences.append("A")

    for subj in subjects:
        name = subj.get("name", "")
        if name in MAIN_SUBJECTS:
            continue
        assigned = subj.get("assigned")
        if assigned is not None and assigned != "":
            assigned_f = safe_float(assigned)
            if assigned_f is not None:
                subject_sum += assigned_f
                confidences.append("B")

    if not confidences:
        return None

    subject_sum = round(subject_sum, 1)
    if subject_sum > FULL_SCORE:
        return None  # 超满分上限，不参与融合

    return {"sum": subject_sum, "confidences": confidences}


def run(data):
    workspace = os.path.abspath(data.get("workspace", "."))

    # 输入校验
    total_score = data.get("total_score")
    if total_score is None:
        return {"status": "error", "reason": "缺少必填字段: total_score"}
    total_score = safe_float(total_score)
    if total_score is None or total_score <= 0:
        return {"status": "error", "reason": f"total_score 值无效: {data.get('total_score')}"}

    # 满分制换算：450分制 → 750分制
    # 始终复制输入 dict，避免修改调用方数据
    data = dict(data)
    try:
        score_scale = safe_int(data.get("score_scale", FULL_SCORE), FULL_SCORE)
        if score_scale is None:
            score_scale = FULL_SCORE
    except (ValueError, TypeError):
        score_scale = FULL_SCORE
    original_total_score = total_score  # 保存原始制总分，供单科比例计算使用
    if score_scale == MAIN_FULL_SCORE:
        data["_original_total_score"] = original_total_score
        data["total_score"] = original_total_score * FULL_SCORE / MAIN_FULL_SCORE
        if data.get("special_line_exam") or data.get("special_line"):
            sl = data.get("special_line_exam") or data.get("special_line")
            data["special_line_exam"] = float(sl) * FULL_SCORE / MAIN_FULL_SCORE

    macro = read_macro_data(workspace)

    if macro is None:
        return {
            "status": "error",
            "reason": "宏观数据_只读.xlsx 不存在，请先完成初始设置",
        }

    methods = []

    # Try methods in priority order: two_module → population_calibration → score_line → school_threshold → school_lookup → school_subject_lookup → percentile → school_estimate
    result = method_two_module(data, macro)
    if result:
        methods.append(result)

    result = method_population_calibration(data, macro)
    if result:
        methods.append(result)

    result = method_score_line(data, macro)
    if result:
        methods.append(result)

    result = method_school_threshold(data, macro)
    if result:
        methods.append(result)

    result = method_school_lookup(data, macro)
    if result:
        methods.append(result)

    result = method_school_subject_lookup(data, macro)
    if result:
        methods.append(result)

    result = method_percentile(data, macro)
    if result:
        methods.append(result)

    result = method_school_estimate(data, macro)
    if result:
        methods.append(result)

    if not methods:
        return {
            "status": "insufficient_data",
            "reason": "当前数据不足以计算等效分。至少需要以下之一：本次考试特控线、全市/联盟排名、校内排名+对照表。",
        }

    # ── 置信度门槛：仅C级方法不可用 ──
    # C级（校排名估算）误差±15分，跨度过大无决策价值
    # 需要至少一个A或B级方法才返回结果
    best_confidence = max(methods, key=lambda m: CONFIDENCE_ORDER.get(m["confidence"], 0))["confidence"]
    if best_confidence in ("C", "D"):
        return {
            "status": "insufficient_data",
            "reason": "当前仅有低精度数据（校排名），等效分误差±15分无参考价值。"
                      "请补充以下任一数据以获得可用结果：1）本次考试特控线（问老师）→A级(±5分)；"
                      "2）全市/联盟排名+总人数 →A级(±5分)。",
        }

    # 报告的置信度取主方法的置信度，与误差区间一致
    primary = methods[0]  # Highest priority method
    equivalent_score = primary["score"]
    primary_confidence = primary["confidence"]
    reported_confidence = primary_confidence if primary_confidence in ("A", "B") else best_confidence
    margin = ERROR_MARGINS.get(primary_confidence, 10)
    error_lower = round(max(0, equivalent_score - margin), 1)
    error_upper = round(min(FULL_SCORE, equivalent_score + margin), 1)

    # Full method details (all methods, for transparency)
    method_details = []
    for m in methods:
        method_details.append({
            "method": m["method"],
            "score": m["score"],
            "confidence": m["confidence"],
            "weight": CONFIDENCE_WEIGHTS.get(m["confidence"], 0),
            "detail": m.get("detail", ""),
        })

    # calculation_detail: primary method as base, fusion appended later if applicable
    calculation_detail = primary.get("detail", "")
    if len(methods) >= 2:
        cv_names = "、".join(m["method"] for m in methods[1:])
        calculation_detail += f"（交叉验证：{cv_names}）"
    calculation_detail = f"[主方法] {calculation_detail}"

    # Cross-validations: supplementary methods vs primary
    cross_validations = []
    for m in methods[1:]:
        diff = round(m["score"] - primary["score"], 1)
        cross_validations.append({
            "method": m["method"],
            "score": m["score"],
            "confidence": m["confidence"],
            "difference": diff,
        })

    # ── 方法分歧处理（三档） ──
    trust_note = None
    divergence = None
    if len(methods) >= 2:
        scores = [m["score"] for m in methods]
        max_diff = max(scores) - min(scores)
        if max_diff <= DIVERGENCE_LOW:
            trust_note = "交叉验证一致，等效分可信度较高"
            divergence = "low"
        elif max_diff <= DIVERGENCE_MEDIUM:
            trust_note = f"方法间存在分歧（最大差异{max_diff:.0f}分），以{primary['method']}为准"
            divergence = "medium"
        else:
            trust_note = f"方法分歧较大（最大差异{max_diff:.0f}分），建议补充排名或特控线数据以提高可靠性"
            divergence = "high"

    # best_confidence already determined above (before C-level gate)

    # ── 数据一致性校验 ──
    warnings = []
    user_total = data.get("city_total") or data.get("alliance_total")
    if user_total:
        user_total = safe_int(user_total)
        if user_total is not None and user_total > 0:
            score_table = filter_score_table(macro.get("一分一段表", []))
            if score_table:
                max_count = max(int(r.get("累计人数", 0)) for r in score_table)
                if max_count > 0 and user_total > 0:
                    ratio = abs(user_total - max_count) / max(max_count, user_total)
                    if ratio > DATA_CONSISTENCY_RATIO:
                        warnings.append(
                            f"考试总人数({user_total})与一分一段表基数({max_count})差异{ratio:.0%}，"
                            "等效分可能存在偏差"
                        )

    # ── 单科等效分（展示用，从总分比例分配，保证各科之和=总分）──
    data["_total_equivalent"] = equivalent_score
    # 延迟计算：融合时仅算一次
    subject_scores = []

    # ── 多方法加权融合 ──
    # 融合公式：所有可用方法 + 单科加总按置信度权重加权平均
    # 单科加总独立计算（语数英用分数线对照法，选科赋分直映），
    # 不与总分法恒等，确保融合产生有意义的交叉校验
    # 单科加总衰减因子 0.5，降低其在融合中的比重
    independent_subj = compute_independent_subject_sum(data, macro)

    components = []  # [(score, weight, label), ...]
    for m in methods:
        w = CONFIDENCE_WEIGHTS.get(m["confidence"], 0)
        if w > 0 and m["confidence"] in ("A", "B"):
            components.append((m["score"], w, m["method"]))

    if independent_subj:
        subj_confs = independent_subj["confidences"]
        subj_weights = [CONFIDENCE_WEIGHTS.get(c, 0) for c in subj_confs]
        w_subject = (sum(subj_weights) / len(subj_weights) * SUBJECT_SUM_DECAY) if subj_weights else 0
        if w_subject > 0:
            components.append((independent_subj["sum"], w_subject, "单科加总"))

    if len(components) >= MIN_DATA_FOR_FUSION:
        weighted_sum = sum(s * w for s, w, _ in components)
        total_weight = sum(w for _, w, _ in components)
        fused = round(weighted_sum / total_weight, 1)

        parts = [f"{label}{score}分(w={w:.2f})" for score, w, label in components]
        calculation_detail += f" | [融合] {' + '.join(parts)} → {fused}分"

        equivalent_score = fused
        # 用融合后的总分重算各科等效分，保证各科加总=总分
        data["_total_equivalent"] = equivalent_score
        subject_scores = compute_subject_equivalents(data, macro)
        # 误差区间基于融合分 ± 最大方法间偏差
        all_scores = [s for s, _, _ in components]
        max_dev = max(abs(fused - s) for s in all_scores)
        error_lower = round(max(0, fused - max(margin, max_dev + 3)), 1)
        error_upper = round(min(FULL_SCORE, fused + max(margin, max_dev + 3)), 1)
    else:
        # 融合不触发时用主方法等效分计算单科
        subject_scores = compute_subject_equivalents(data, macro)

    result = {
        "status": "ok",
        "primary_method": primary["method"],
        "equivalent_score": equivalent_score,
        "confidence": reported_confidence,
        "error_lower": error_lower,
        "error_upper": error_upper,
        "calculation_detail": calculation_detail,
        "method_count": len(methods),
        "method_details": method_details,
        "cross_validations": cross_validations,
        "trust_note": trust_note,
        "divergence": divergence,
        "warnings": warnings,
        "subject_scores": subject_scores,
    }
    return result


def main():
    try:
        data = json.loads(sys.stdin.read())
        result = run(data)
    except Exception as e:
        result = {"status": "error", "reason": f"计算过程异常: {e}"}
    print(json.dumps(result, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
