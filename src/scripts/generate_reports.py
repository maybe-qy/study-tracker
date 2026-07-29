#!/usr/bin/env python3
"""Generate 8 HTML reports from Excel data.

Reports:
  1. 个人档案.html — latest equivalent score, status, target gap
  2. 高考总分趋势.html — equivalent score time series + analysis
  3-8. [语文/数学/英语/选1/选2/选3]追踪.html — single subject tracking

Usage:
  python generate_reports.py --workspace <path>
"""

import argparse
import base64
import json
import os
import sys
import statistics
from datetime import datetime

from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader

from excel_utils import read_sheet_dicts, read_macro_data
from config import *  # noqa: F401,F403 — 统一导入 CONFIDENCE_WEIGHTS / EWMA_ALPHA / TREND_SLOPE_THRESHOLD 等常量


def safe_float(val, default=None):
    """安全转换为 float，失败时返回 default。排除 bool 类型。"""
    if isinstance(val, bool):
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default

DISCLAIMER = """声明与局限性

1. 等效分方法：
   优先使用双模块换算法（各科校内划线独立换算），或分数线对照法（省级特控线固定锚点）。
   校内排名对照法（有本校高考对照表时）为 A 级。
   全市/联盟排名锚定法作为交叉验证。
   校排名估算为 C 级（低精度回退），仅C级可用时返回insufficient_data。
   等效分仅供参考，不构成对高考成绩的预测。

2. 置信度分级：
   A级：双模块换算法（多数模块）、分数线对照法、校内排名对照法、单科排名对照法、全市/联盟排名锚定法、全市/联盟统一赋分。
   B级：双模块换算法（部分模块）、主科原始分、全市统考/联盟考试中无独立划线的选科。
   C级：校排名估算（无本校高考对照数据）。
   D级：无排名无分数线分数。
   趋势/波动分析权重：A=1.0, B=0.8, C=0.5, D 不参与。

3. 等效分使用赋分：
   总分 = 语数英原始分 + 选科赋分（有赋分用赋分，无赋分用原始分）。
   单科等效分优先用赋分数据计算。

4. 数据来源：用户上传。"""


def load_data(workspace):
    """Load all Excel data."""
    data = {"exams": [], "equivalent": [], "subjects": {}, "volatility": []}

    def sort_by_date(records, reverse=False):
        """按日期排序，默认升序（最旧的在前，便于趋势分析）"""
        from datetime import datetime as _dt
        def parse_date(d):
            if not d or not isinstance(d, str):
                return _dt.min
            d = d.strip()
            # 尝试多种日期格式，补零标准化
            for fmt in ("%Y-%m-%d", "%Y-%m", "%Y/%m/%d", "%Y/%m"):
                try:
                    return _dt.strptime(d, fmt)
                except ValueError:
                    continue
            # 尝试补零: "2026-1-5" → "2026-01-05"
            try:
                parts = d.replace("/", "-").split("-")
                if len(parts) >= 2:
                    parts = [p.zfill(2) if len(p) < 2 else p for p in parts]
                    return _dt.strptime("-".join(parts), "%Y-%m-%d")
            except ValueError:
                pass
            return _dt.min
        return sorted(records, key=lambda r: parse_date(r.get("日期", "")), reverse=reverse)

    # 成绩总表
    path = os.path.join(workspace, "data", "personal", "成绩总表.xlsx")
    if os.path.exists(path):
        wb = load_workbook(path, data_only=True)
        try:
            if "成绩总表" in wb.sheetnames:
                data["exams"] = sort_by_date(read_sheet_dicts(wb["成绩总表"]))
        finally:
            wb.close()

    # 等效分记录
    path = os.path.join(workspace, "data", "personal", "等效分记录.xlsx")
    if os.path.exists(path):
        wb = load_workbook(path, data_only=True)
        try:
            if "等效分记录" in wb.sheetnames:
                data["equivalent"] = sort_by_date(read_sheet_dicts(wb["等效分记录"]))
        finally:
            wb.close()

    # 单科追踪
    path = os.path.join(workspace, "data", "personal", "单科追踪.xlsx")
    if os.path.exists(path):
        wb = load_workbook(path, data_only=True)
        try:
            for name in wb.sheetnames:
                data["subjects"][name] = sort_by_date(read_sheet_dicts(wb[name]))
        finally:
            wb.close()

    # 宏观数据
    macro_data = read_macro_data(workspace)
    data["macro"] = macro_data if macro_data is not None else {}

    # 学校招生数据
    path = os.path.join(workspace, "data", "school", "学校招生_只读.xlsx")
    data["admission"] = {}
    if os.path.exists(path):
        wb = load_workbook(path, data_only=True)
        try:
            for name in wb.sheetnames:
                data["admission"][name] = read_sheet_dicts(wb[name])
        finally:
            wb.close()

    return data


# ─── HTML generation helpers ──────────────────────────────────────────
# CONFIDENCE_WEIGHTS is imported from config.py (via `from config import *`)


def ewma(scores, alpha=EWMA_ALPHA):
    """指数加权移动平均（Exponentially Weighted Moving Average）。

    共享实现，消除 4 处重复：render_personal / prediction_state /
    eval_labels / render_subject。alpha 默认取 config.EWMA_ALPHA=0.3。
    """
    if not scores:
        return 0
    result = scores[0]
    for s in scores[1:]:
        result = alpha * s + (1 - alpha) * result
    return result


def parse_eq_detail(detail_str):
    """解析等效分记录的"详细信息" JSON，返回 dict 或 None。

    共享实现，消除 3 处重复：render_personal / render_trend / render_subject。
    对 calculation_detail 与 subject_scores 做防御性类型转换：
      - calculation_detail：list → "|".join；非 str → str()
      - subject_scores：dict → list-of-dict；非 list → []
    """
    if not detail_str:
        return None
    try:
        obj = json.loads(detail_str)
        # 防御性转换
        cd = obj.get("calculation_detail", "")
        if isinstance(cd, list):
            cd = "|".join(str(x) for x in cd)
        elif not isinstance(cd, str):
            cd = str(cd)
        obj["calculation_detail"] = cd
        ss = obj.get("subject_scores", [])
        if isinstance(ss, dict):
            ss = [{"subject": k, "score": v} for k, v in ss.items()]
        elif not isinstance(ss, list):
            ss = []
        obj["subject_scores"] = ss
        return obj
    except (json.JSONDecodeError, TypeError):
        return None


def filter_weighted(records):
    """Extract (score, weight) tuples from equivalent score records, excluding D-level."""
    weighted = []
    for r in records:
        conf = str(r.get("置信度", "A")).strip().replace("级", "").upper()
        weight = CONFIDENCE_WEIGHTS.get(conf, 1.0)
        if weight > 0:
            score_val = r.get("等效分（融合结果）")
            if score_val is not None and score_val != "":
                score = safe_float(score_val)
                if score is not None:
                    weighted.append((score, weight))
    return weighted


def extract_eq_scores(eq_records):
    """从等效分记录中提取有效分数列表（排除 None/空/非数值）。"""
    scores = []
    for r in eq_records:
        val = r.get("等效分（融合结果）")
        if val is not None and val != "":
            score = safe_float(val)
            if score is not None:
                scores.append(score)
    return scores


def compute_trend(scores):
    """Determine trend direction: 'up', 'down', or 'flat'. Returns (class, arrow, text)."""
    if len(scores) < 2:
        return ("flat", "→", "数据不足")
    # Simple linear trend on last MIN_DATA_FOR_ANALYSIS or all
    recent = scores[-MIN_DATA_FOR_ANALYSIS:] if len(scores) >= MIN_DATA_FOR_ANALYSIS else scores
    n = len(recent)
    if n < 2:
        return ("flat", "→", "持平")
    # Slope of best-fit line
    x_mean = (n - 1) / 2
    y_mean = sum(recent) / n
    num = sum((i - x_mean) * (recent[i] - y_mean) for i in range(n))
    den = sum((i - x_mean) ** 2 for i in range(n))
    if den == 0:
        return ("flat", "→", "持平")
    slope = num / den
    if slope > TREND_SLOPE_THRESHOLD:
        return ("up", "↑", "上升")
    elif slope < -TREND_SLOPE_THRESHOLD:
        return ("down", "↓", "下降")
    else:
        return ("flat", "→", "持平")


def compute_volatility(scores):
    """Compute sigma and volatility range. Returns (sigma, lower, upper)."""
    if len(scores) < MIN_DATA_FOR_ANALYSIS:
        return (None, None, None)
    mean = statistics.mean(scores)
    sigma = statistics.stdev(scores)
    return (round(sigma, 1), round(mean - VOLATILITY_SIGMA_MULT * sigma, 1), round(mean + VOLATILITY_SIGMA_MULT * sigma, 1))


def compute_volatility_weighted(weighted_scores):
    """Weighted sigma and volatility range."""
    if len(weighted_scores) < MIN_DATA_FOR_ANALYSIS:
        return (None, None, None)
    scores = [s for s, _ in weighted_scores]
    weights = [w for _, w in weighted_scores]
    w_mean = sum(s * w for s, w in zip(scores, weights)) / sum(weights)
    w_var = sum(w * (s - w_mean) ** 2 for s, w in zip(scores, weights)) / sum(weights)
    sigma = w_var ** 0.5
    return (round(sigma, 1), round(w_mean - VOLATILITY_SIGMA_MULT * sigma, 1), round(w_mean + VOLATILITY_SIGMA_MULT * sigma, 1))


def prediction_state(scores):
    """Compute prediction label for latest score. Returns '积极'/'正常'/'消极'."""
    if len(scores) < MIN_DATA_FOR_ANALYSIS:
        return None
    # HP-filter simplified: use EWMA trend
    # residual[i] = scores[i] - ewma(scores[:i])（基于前 i 个点的 EWMA 基线）
    residuals = [scores[i] - ewma(scores[:i]) for i in range(1, len(scores))]
    if len(residuals) < 2:
        return "正常"
    q75 = sorted(residuals)[int(len(residuals) * 0.75)]
    q25 = sorted(residuals)[int(len(residuals) * 0.25)]
    latest = residuals[-1]
    if latest >= q75:
        return "积极"
    elif latest <= q25:
        return "消极"
    else:
        return "正常"


def eval_labels(scores):
    """Count positive/normal/negative labels + return label sequence for trend detection."""
    if len(scores) < MIN_DATA_FOR_ANALYSIS:
        return (None, None)
    labels = {"积极": 0, "正常": 0, "消极": 0}
    sequence = []
    # 预热：用前3个点建立EWMA基线（ewma(scores[:3]) 隐式完成）
    # 从第4个点开始标注（EWMA基于前i个点，当前点用于比较）
    for i in range(3, len(scores)):
        baseline = ewma(scores[:i])
        if scores[i] > baseline + PREDICTION_THRESHOLD:
            labels["积极"] += 1
            sequence.append("积极")
        elif scores[i] < baseline - PREDICTION_THRESHOLD:
            labels["消极"] += 1
            sequence.append("消极")
        else:
            labels["正常"] += 1
            sequence.append("正常")
    return (labels, sequence)


def classify_volatility_style(labels, sigma, sequence):
    """Classify volatility pattern. Returns descriptive label or None if insufficient data."""
    if labels is None or sequence is None or sigma is None:
        return None
    total = labels["积极"] + labels["正常"] + labels["消极"]
    if total == 0:
        return None
    normal_ratio = labels["正常"] / total
    active_ratio = (labels["积极"] + labels["消极"]) / total
    # 趋势型: 3+ consecutive same direction (excluding "正常")
    max_consecutive = 1
    current_run = 1
    for i in range(1, len(sequence)):
        if sequence[i] == sequence[i-1] and sequence[i] != "正常":
            current_run += 1
            max_consecutive = max(max_consecutive, current_run)
        else:
            current_run = 1
    if max_consecutive >= 3:
        return "呈持续变化趋势"
    if active_ratio >= 0.5:
        return "分数波动较大"
    if normal_ratio >= 0.7:
        return "分数相对稳定"
    return "分数波动较大"


def _find_logo_base64(target_university, logo_dir=None):
    """Find university logo SVG and return as base64 data URI, or None if not found."""
    if not target_university:
        return None
    if logo_dir is None:
        logo_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "logos")
    if not os.path.isdir(logo_dir):
        return None
    for fname in os.listdir(logo_dir):
        if target_university in fname and fname.endswith(".svg"):
            fpath = os.path.join(logo_dir, fname)
            try:
                with open(fpath, "rb") as fh:
                    return "data:image/svg+xml;base64," + base64.b64encode(fh.read()).decode()
            except Exception:
                continue  # 跳过读取失败的文件，继续查找
    return None


# ─── Report generators ─────────────────────────────────────────────────

def _compute_tier_info(macro, score, target_university, target_line, target_gap):
    """Compute university tier information for the personal report.

    Extracts tier matching logic from render_personal() for readability.
    Returns tier_info dict or None.
    """
    tier_info = None
    tier_data = macro.get("院校层次", [])
    if tier_data and score > 0:
        current_tier = None
        next_tier = None
        all_tiers = []

        for row in tier_data:
            scope = str(row.get("范围", ""))
            name = str(row.get("梯队", ""))
            threshold_str = str(row.get("预估总分门槛", "0"))
            upper_str = str(row.get("预估总分上限", "750"))
            try:
                threshold = float(threshold_str)
                upper = float(upper_str) if upper_str else FULL_SCORE
            except (ValueError, TypeError):
                continue

            tier_entry = {
                "scope": scope,
                "name": name,
                "threshold": threshold,
                "upper": upper,
                "schools": str(row.get("代表院校", "")),
                "is_current": False,
            }

            if threshold <= score <= upper:
                tier_entry["is_current"] = True
                current_tier = tier_entry

            all_tiers.append(tier_entry)

        if current_tier:
            above = [t for t in all_tiers if t["threshold"] > current_tier["upper"]]
            above.sort(key=lambda t: t["threshold"])
            if above:
                next_tier = above[0]
            elif [t for t in all_tiers if t["threshold"] > score]:
                candidates = [t for t in all_tiers if t["threshold"] > score]
                candidates.sort(key=lambda t: t["threshold"])
                next_tier = candidates[0]

        tier_info = {
            "current": current_tier,
            "next": next_tier,
            "next_gap": round(next_tier["threshold"] - score, 0) if next_tier else None,
            "all_tiers": all_tiers,
            "target_university": target_university,
            "target_line": target_line,
            "target_gap": target_gap,
            "target_logo": _find_logo_base64(target_university),
        }

    # 无院校层次参考但有目标院校时，构建最小 tier_info
    if tier_info is None and target_university:
        tier_info = {
            "current": None,
            "next": None,
            "next_gap": None,
            "all_tiers": [],
            "target_university": target_university,
            "target_line": target_line,
            "target_gap": target_gap,
            "target_logo": _find_logo_base64(target_university),
        }

    return tier_info


def render_personal(data, env):
    """Render 个人档案.html."""
    eq_records = data["equivalent"]
    macro = data.get("macro", {})

    exam_records = data.get("exams", [])
    if not eq_records:
        has_exams = len(exam_records) >= 1
        template = env.get_template("report_personal.html")
        return template.render(
            generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
            equivalent_score="暂无数据" if not has_exams else "等待计算",
            latest_equiv=0,
            confidence="-",
            method="-",
            calc_detail="",
            error_lower="-",
            error_upper="-",
            has_analysis=False,
            is_first_record=False,  # 个人档案的首次引导仅在已有1条等效分时触发
            exam_count=len(exam_records),
            trend_class="flat",
            trend_arrow="→",
            trend_text="等待数据",
            prediction_state="-",
            volatility_lower="-",
            volatility_upper="-",
            sigma="-",
            subject_scores=[],
            tier_info=None,
            volatility_style="-",
            disclaimer=DISCLAIMER,
        )

    latest = eq_records[-1]
    latest_equiv = safe_float(latest.get("等效分（融合结果）"))
    eq_scores = extract_eq_scores(eq_records)
    weighted = filter_weighted(eq_records)

    trend_class, trend_arrow, trend_text = compute_trend(eq_scores)
    sigma, vol_low, vol_high = compute_volatility_weighted(weighted)
    pred = prediction_state(eq_scores)
    has_analysis = len(eq_scores) >= MIN_DATA_FOR_ANALYSIS
    is_first_record = len(eq_scores) == 1
    labels, label_sequence = eval_labels(eq_scores) if len(eq_scores) >= MIN_DATA_FOR_ANALYSIS else (None, None)
    volatility_style = classify_volatility_style(labels, sigma, label_sequence) if has_analysis else None

    # ── 院校定位 ──
    # 目标院校始终从 latest 提取，不依赖院校层次参考数据
    target_university = latest.get("目标院校")
    target_line = safe_float(latest.get("目标院校录取线"))
    target_gap = safe_float(latest.get("差距分数"))
    # Auto-compute gap if not stored but we have both values
    score = latest_equiv if latest_equiv is not None else (eq_scores[-1] if eq_scores else 0)
    if target_gap is None and target_line is not None and score > 0:
        target_gap = round(score - target_line, 1)

    tier_info = _compute_tier_info(macro, score, target_university, target_line, target_gap)

    # Extract calculation detail from latest record
    latest_calc_detail = ""
    latest_subject_scores = []
    detail_obj = parse_eq_detail(latest.get("详细信息", ""))
    if detail_obj:
        latest_calc_detail = detail_obj.get("calculation_detail", "")
        latest_subject_scores = detail_obj.get("subject_scores", [])

    template = env.get_template("report_personal.html")
    return template.render(
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
        equivalent_score=f"{latest_equiv:.0f} 分" if latest_equiv is not None else "暂无",
        latest_equiv=latest_equiv if latest_equiv is not None else 0,
        confidence=str(latest.get("置信度", "-")).replace("级", "").strip() or "-",
        method=latest.get("主计算方法", "-"),
        calc_detail=latest_calc_detail,
        error_lower=latest.get("误差区间下限", "-"),
        error_upper=latest.get("误差区间上限", "-"),
        has_analysis=has_analysis,
        trend_class=trend_class,
        trend_arrow=trend_arrow,
        trend_text=trend_text,
        prediction_state=pred or "-",
        volatility_lower=vol_low if vol_low is not None else "-",
        volatility_upper=vol_high if vol_high is not None else "-",
        sigma=f"{sigma}分" if sigma is not None else "-",
        volatility_style=volatility_style or "-",
        is_first_record=is_first_record,
        exam_count=len(eq_scores),
        subject_scores=latest_subject_scores,
        tier_info=tier_info,
        disclaimer=DISCLAIMER,
    )


def render_trend(data, env):
    """Render 高考总分趋势.html."""
    eq_records = data["equivalent"]
    exam_records = data.get("exams", [])

    if not eq_records:
        # 有考试记录但无等效分时，显示首次录入引导
        is_first = len(exam_records) >= 1
        exams_for_display = []
        if is_first:
            for e in exam_records:
                exams_for_display.append({
                    "date": e.get("日期", "-"),
                    "name": e.get("考试名", "-"),
                    "score": "-",
                    "confidence": "-",
                    "method": "等待计算",
                    "method_switch": False,
                    "calc_detail": "",
                    "prev_method": "",
                })
            exams_for_display = list(reversed(exams_for_display))
        template = env.get_template("report_trend.html")
        return template.render(
            generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
            exams=exams_for_display,
            has_analysis=False,
            trend_class="flat",
            trend_arrow="→",
            trend_text="等待数据",
            sigma="-",
            volatility_lower="-",
            volatility_upper="-",
            labels={"positive": "-", "normal": "-", "negative": "-"},
            cross_validations=[],
            volatility_style="-",
            is_first_record=is_first,
            exam_count=len(exam_records),
            disclaimer=DISCLAIMER,
        )

    exams = []
    for r in eq_records:
        # Extract calculation detail from 详细信息 JSON
        detail_obj = parse_eq_detail(r.get("详细信息", ""))
        calc_detail = detail_obj.get("calculation_detail", "") if detail_obj else ""

        exams.append({
            "date": r.get("日期", "-"),
            "name": r.get("考试名", "-"),
            "score": r.get("等效分（融合结果）", "-"),
            "confidence": str(r.get("置信度", "-")).replace("级", "").strip() or "-",
            "method": r.get("主计算方法", "-"),
            "calc_detail": calc_detail,
            "method_switch": False,  # will be set below
            "prev_method": "",  # initialized for consistent dict structure
        })

    # I15: 检测方法切换，标记相邻两次考试方法不同的记录
    for i in range(1, len(exams)):
        if exams[i].get("method") != exams[i-1].get("method"):
            exams[i]["method_switch"] = True
            exams[i]["prev_method"] = exams[i-1].get("method", "")

    # 显示时反转为降序（最新的在最上面，方便查看近期趋势）
    exams = list(reversed(exams))

    eq_scores = extract_eq_scores(eq_records)
    weighted = filter_weighted(eq_records)
    trend_class, trend_arrow, trend_text = compute_trend(eq_scores)
    sigma, vol_low, vol_high = compute_volatility_weighted(weighted)
    has_analysis = len(eq_scores) >= MIN_DATA_FOR_ANALYSIS
    is_first_record = len(eq_scores) == 1
    labels, label_sequence = eval_labels(eq_scores) if len(eq_scores) >= MIN_DATA_FOR_ANALYSIS else (None, None)
    volatility_style = classify_volatility_style(labels, sigma, label_sequence) if has_analysis else None

    # Cross validations summary — extract both method 1 and method 2
    cross_validations = []
    for r in eq_records:
        for cv_num in ("1", "2"):
            cv_method = r.get(f"交叉验证方法{cv_num}")
            cv_score = r.get(f"交叉验证分{cv_num}")
            if cv_method and cv_score:
                diff = None
                primary = safe_float(r.get("等效分（融合结果）"))
                cv_score_f = safe_float(cv_score)
                if cv_score_f is None:
                    continue
                if primary is not None:
                    diff = f"{cv_score_f - primary:+.1f}"
                cross_validations.append({
                    "exam": r.get("考试名", "-"),
                    "method": cv_method,
                    "score": cv_score,
                    "diff": diff or "-",
                })

    template = env.get_template("report_trend.html")
    return template.render(
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
        exams=exams,  # 已按日期降序排列（最新在前，便于查看）
        has_analysis=has_analysis,
        trend_class=trend_class,
        trend_arrow=trend_arrow,
        trend_text=trend_text,
        sigma=f"{sigma}" if sigma is not None else "-",
        volatility_lower=vol_low if vol_low is not None else "-",
        volatility_upper=vol_high if vol_high is not None else "-",
        labels={"positive": labels["积极"] if labels else "-", "normal": labels["正常"] if labels else "-", "negative": labels["消极"] if labels else "-"},
        volatility_style=volatility_style or "-",
        is_first_record=is_first_record,
        exam_count=len(exams),
        cross_validations=cross_validations,
        disclaimer=DISCLAIMER,
    )


def _build_subject_record(date, exam, raw, assigned, confidence, subject_name):
    """Build a single subject tracking record dict from raw/assigned scores.

    Shared by 单科追踪 fallback and 成绩总表 fallback in render_subject().
    For 语数英: always uses raw score. For 选科: uses assigned if available.
    """
    is_main = subject_name in ("语文", "数学", "英语")
    use_assigned = assigned is not None and assigned != "" and not is_main

    if use_assigned:
        score_float = safe_float(assigned)
        score_str = f"{score_float:.1f}" if score_float is not None else "-"
    elif raw is not None and raw != "":
        score_float = safe_float(raw)
        score_str = f"{score_float:.1f}" if score_float is not None else "-"
    else:
        score_float = None
        score_str = "-"

    return {
        "date": date or "-",
        "exam": exam or "-",
        "raw": raw if raw is not None and raw != "" else "-",
        "assigned": assigned if assigned is not None and assigned != "" else "-",
        "score": score_str,
        "confidence": confidence or "-",
        "method": "-",
    }, score_float


def render_subject(data, env, subject_name, sheet_name):
    """Render a single subject tracking HTML.
    Reads per-subject equivalent scores from 等效分记录 first;
    falls back to 成绩总表.xlsx exam records.
    """
    eq_records = data.get("equivalent", [])
    records = []
    scores = []

    # Primary: extract per-subject equivalent scores from saved eq data
    if eq_records:
        for eq in eq_records:
            detail_obj = parse_eq_detail(eq.get("详细信息", ""))
            if not detail_obj:
                continue
            for s in detail_obj.get("subject_scores", []):
                if s.get("subject") != subject_name:
                    continue
                score = s.get("score")
                if score is not None:
                    score_f = safe_float(score)
                    if score_f is None:
                        continue
                    scores.append(score_f)
                    records.append({
                        "date": eq.get("日期", "-"),
                        "exam": eq.get("考试名", "-"),
                        "score": f"{score_f:.1f}",
                        "confidence": s.get("confidence", "-"),
                        "method": s.get("method", "-"),
                    })

    # Fallback: extract from 单科追踪.xlsx or 成绩总表
    if not records:
        subject_data = data["subjects"].get(sheet_name, [])
        if subject_data:
            for r in subject_data:
                raw = r.get("原始分")
                assigned = r.get("赋分")
                rec, score_float = _build_subject_record(
                    r.get("日期"), r.get("考试名"), raw, assigned,
                    r.get("赋分置信度"), subject_name,
                )
                records.append(rec)
                if score_float is not None:
                    scores.append(score_float)
        else:
            for exam in data.get("exams", []):
                raw = None
                assigned = None
                conf = None
                if subject_name in ("语文", "数学", "英语"):
                    raw = exam.get(subject_name)
                    conf = "B"
                else:
                    for i in range(1, 4):
                        if str(exam.get(f"选科{i}名称", "")) == subject_name:
                            raw = exam.get(f"选科{i}原始分")
                            assigned = exam.get(f"选科{i}赋分")
                            conf = exam.get(f"选科{i}赋分置信度") or "A"
                            break
                if raw is None or raw == "":
                    continue
                rec, score_float = _build_subject_record(
                    exam.get("日期"), exam.get("考试名"), raw, assigned,
                    conf, subject_name,
                )
                records.append(rec)
                if score_float is not None:
                    scores.append(score_float)

    valid_scores = [s for s in scores if s is not None]
    if not valid_scores:
        # No data for this subject, still render an empty report
        dynamic = "-"
        latest = "-"
        highest = "-"
        trend_class, trend_arrow, trend_text = "flat", "→", "无数据"
    else:
        # EWMA for dynamic score (α=EWMA_ALPHA=0.3，越近权重越高)
        dynamic = round(ewma(valid_scores, alpha=EWMA_ALPHA), 1)
        latest = valid_scores[-1]
        highest = max(valid_scores)
        trend_class, trend_arrow, trend_text = compute_trend(valid_scores)

    is_first_record = len(valid_scores) == 1
    template = env.get_template("report_subject.html")
    return template.render(
        subject=subject_name,
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
        dynamic_score=dynamic,
        latest=latest,
        highest=highest,
        trend_class=trend_class,
        trend_arrow=trend_arrow,
        trend_text=trend_text,
        is_first_record=is_first_record,
        exam_count=len(valid_scores),
        records=list(reversed(records)),  # 降序显示（最新在前）
        disclaimer=DISCLAIMER,
    )


def run(workspace):
    data = load_data(workspace)
    data["_workspace"] = workspace

    # Ensure output directory exists
    output_dir = os.path.join(workspace, "output")
    os.makedirs(output_dir, exist_ok=True)

    # Setup Jinja2
    assets_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "assets")
    if not os.path.isdir(assets_dir):
        print(json.dumps({"status": "error", "reason": f"模板目录不存在: {assets_dir}"}))
        sys.exit(1)

    env = Environment(loader=FileSystemLoader(assets_dir))
    generated = []
    errors = []

    # 每个报告独立渲染，单个失败不影响其他报告
    def _safe_render(render_fn, *args, **kwargs):
        """安全渲染单个报告，异常时记录错误而非中断整个流程。"""
        try:
            return render_fn(*args, **kwargs)
        except Exception as e:
            errors.append(f"{render_fn.__name__}: {e}")
            return None

    # 1. 个人档案
    html = _safe_render(render_personal, data, env)
    if html:
        p = os.path.join(workspace, "output", "个人档案.html")
        with open(p, "w", encoding="utf-8") as f:
            f.write(html)
        generated.append(p)

    # 2. 高考总分趋势
    html = _safe_render(render_trend, data, env)
    if html:
        p = os.path.join(workspace, "output", "高考总分趋势.html")
        with open(p, "w", encoding="utf-8") as f:
            f.write(html)
        generated.append(p)

    # 3-8. 单科追踪 x6
    # Determine subject names from exam data
    subject_sheet_map = {
        "语文追踪": "语文",
        "数学追踪": "数学",
        "英语追踪": "英语",
        "选科1追踪": "选科1",
        "选科2追踪": "选科2",
        "选科3追踪": "选科3",
    }

    # Try to get actual subject names from exam data
    exams = data["exams"]
    latest_exam = exams[-1] if exams else None
    if latest_exam:
        sub1_name = latest_exam.get("选科1名称")
        sub2_name = latest_exam.get("选科2名称")
        sub3_name = latest_exam.get("选科3名称")
        if sub1_name:
            subject_sheet_map["选科1追踪"] = str(sub1_name)
        if sub2_name:
            subject_sheet_map["选科2追踪"] = str(sub2_name)
        if sub3_name:
            subject_sheet_map["选科3追踪"] = str(sub3_name)

    for sheet_name, subject_name in subject_sheet_map.items():
        html = _safe_render(render_subject, data, env, subject_name, sheet_name)
        if html:
            p = os.path.join(workspace, "output", f"{subject_name}追踪.html")
            with open(p, "w", encoding="utf-8") as f:
                f.write(html)
            generated.append(p)

    result = {"status": "ok", "files": generated}
    if errors:
        result["warnings"] = errors
    return result


def main():
    parser = argparse.ArgumentParser(description="Generate study-tracker HTML reports")
    parser.add_argument("--workspace", required=True, help="Workspace root path")
    args = parser.parse_args()

    workspace = os.path.abspath(args.workspace)
    if not os.path.isdir(workspace):
        print(json.dumps({"status": "error", "reason": f"路径不是目录: {workspace}"}))
        sys.exit(1)

    result = run(workspace)
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
