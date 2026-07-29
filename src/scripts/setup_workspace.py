#!/usr/bin/env python3
"""Initialize workspace directory tree and empty Excel files with headers.

Usage:
  python setup_workspace.py --workspace <path>
  python setup_workspace.py --workspace "d:/研究/学升"

Idempotent: skips files/dirs that already exist.
Never overwrites files that contain data (row count > 1).
"""

import argparse
import json
import os
import sys

from openpyxl import load_workbook, Workbook

HEADERS = {
    "成绩总表": [
        "考试名", "日期", "类型", "年级",
        "语文", "数学", "英语",
        "选科1名称", "选科1原始分", "选科1赋分", "选科1赋分置信度",
        "选科2名称", "选科2原始分", "选科2赋分", "选科2赋分置信度",
        "选科3名称", "选科3原始分", "选科3赋分", "选科3赋分置信度",
        "总分",
        "市/联盟排名", "市/联盟总人数",
        "校排名", "校总人数",
        "特控线", "重点班未参考人数", "优划线", "满分制", "学校类型", "排名类型", "备注",
    ],
    "等效分记录": [
        "考试名", "日期",
        "等效分（融合结果）", "置信度", "主计算方法",
        "交叉验证方法1", "交叉验证分1",
        "交叉验证方法2", "交叉验证分2",
        "误差区间下限", "误差区间上限",
        "目标院校", "目标院校录取线", "差距分数", "详细信息",
    ],
}

SUBJECT_SHEETS = [
    "语文追踪", "数学追踪", "英语追踪",
    "选科1追踪", "选科2追踪", "选科3追踪",
]

SUBJECT_HEADERS = [
    "考试名", "日期", "原始分", "赋分", "赋分置信度",
]

MACRO_SHEETS = {
    "一分一段表": ["分数", "累计人数", "省份", "年份"],
    "特控线": ["年份", "省份", "特控线分数"],
    "赋分区间": ["省份", "等级", "最低分", "最高分"],
    "本校对照表_总分": ["校内排名", "高考总分"],
    # 院校层次表（generate_reports 读取此 Sheet 生成院校定位）
    "院校层次": ["范围", "梯队", "预估总分门槛", "预估总分上限", "代表院校"],
    # 院校录取线（独立 Sheet，不与院校层次混淆）
    "院校层次_录取线": ["院校名称", "年份", "录取最低分", "录取最低位次"],
    # 校内划线数据（用户补充，启用双模块换算法和校排阈值估算法）
    # 非标准布局：需包含"特控线分段"和"浙大线分段"标记行
    "校内划线_升级": [
        "特控线分段",
        "科目", "2027划线", "2027上线", "2028划线", "2028上线",
        "浙大线分段",
        "科目", "2027划线", "2027上线", "2028划线", "2028上线",
    ],
    # 门槛数据（用户补充，启用人数校准法）
    "门槛数据": ["考试名称", "特控线分数", "特控线上线人数", "浙大线分数", "浙大线上线人数"],
    # 单科对照表（用户补充，启用单科排名对照法）
    "单科对照": ["科目", "校内排名", "高考等效分"],
}

SCHOOL_SHEETS = {
    "深大AI录取数据": ["年份", "专业", "录取最低分", "录取最低位次"],
    "浙大2026投档线": ["专业名称", "投档线", "位次"],
    "2026浙江一段投档线": ["院校名称", "专业名称", "投档线", "位次"],
}


def create_dirs(workspace):
    dirs = [
        os.path.join(workspace, "data", "macro"),
        os.path.join(workspace, "data", "school"),
        os.path.join(workspace, "data", "personal", "individual"),
    ]
    created = []
    for d in dirs:
        if not os.path.exists(d):
            os.makedirs(d, exist_ok=True)
            created.append(d)
    return created


def create_excel(path, sheet_headers):
    """Create an Excel file with headers. sheet_headers is {sheet_name: [headers]}.

    Idempotency: skip if file exists AND has data (row count > 1).
    Recreate if file exists but only has headers (row count == 1 or file is empty).
    """
    if os.path.exists(path):
        # Check if file has actual data (not just headers)
        try:
            existing = load_workbook(path, data_only=True, read_only=True)
            has_data = any(ws.max_row > 1 for ws in existing.worksheets)
            existing.close()
            if has_data:
                return None  # Don't overwrite existing data
            # Else: file exists but only has headers, will recreate
        except Exception:
            # If file is corrupted, try to recreate
            pass

    wb = Workbook()
    try:
        first = True
        for sheet_name, headers in sheet_headers.items():
            if first:
                ws = wb.active
                ws.title = sheet_name
                first = False
            else:
                ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)
        wb.save(path)
    finally:
        wb.close()
    return path


def run(workspace):
    result = {"status": "ok", "created": [], "skipped": [], "errors": []}

    # 1. Create directories
    result["created"].extend(create_dirs(workspace))

    # 2. Create 成绩总表.xlsx
    p = os.path.join(workspace, "data", "personal", "成绩总表.xlsx")
    created = create_excel(p, {"成绩总表": HEADERS["成绩总表"]})
    if created:
        result["created"].append(created)
    else:
        result["skipped"].append(p)

    # 3. Create 等效分记录.xlsx
    p = os.path.join(workspace, "data", "personal", "等效分记录.xlsx")
    created = create_excel(p, {"等效分记录": HEADERS["等效分记录"]})
    if created:
        result["created"].append(created)
    else:
        result["skipped"].append(p)

    # 4. Create 单科追踪.xlsx with 6 sheets
    p = os.path.join(workspace, "data", "personal", "单科追踪.xlsx")
    created = create_excel(p, {s: SUBJECT_HEADERS for s in SUBJECT_SHEETS})
    if created:
        result["created"].append(created)
    else:
        result["skipped"].append(p)

    # 5. Create 宏观数据_只读.xlsx
    p = os.path.join(workspace, "data", "macro", "宏观数据_只读.xlsx")
    created = create_excel(p, MACRO_SHEETS)
    if created:
        result["created"].append(created)
    else:
        result["skipped"].append(p)

    # 6. Create 学校招生_只读.xlsx
    p = os.path.join(workspace, "data", "school", "学校招生_只读.xlsx")
    created = create_excel(p, SCHOOL_SHEETS)
    if created:
        result["created"].append(created)
    else:
        result["skipped"].append(p)

    return result


def main():
    parser = argparse.ArgumentParser(description="Setup workspace for study-tracker")
    parser.add_argument("--workspace", required=True, help="Workspace root path")
    args = parser.parse_args()

    workspace = os.path.abspath(args.workspace)
    if not os.path.exists(workspace):
        print(json.dumps({"status": "error", "reason": f"路径不存在: {workspace}"}))
        sys.exit(1)

    result = run(workspace)
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
