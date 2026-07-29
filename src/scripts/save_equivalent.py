#!/usr/bin/env python3
"""Save equivalent score calculation results to 等效分记录.xlsx.

Usage:
  python calc_equivalent.py < exam_data.json | python save_equivalent.py --workspace <path>

  Or combined:
  echo '{"workspace":".", "exam_name":"...", ...}' | python calc_equivalent.py | \
    python save_equivalent.py --workspace . --exam-name "..." --exam-date "..." --target "..." --target-line 652

Input JSON fields:
  workspace, exam_name, exam_date — required
  target_university, target_line — optional
  Plus the calc_equivalent output via stdin.
"""

import argparse
import json
import os
import sys
from openpyxl import load_workbook


def run(workspace, exam_name, exam_date, calc_result, target_university=None, target_line=None):
    path = os.path.join(workspace, "data", "personal", "等效分记录.xlsx")
    if not os.path.exists(path):
        return {"status": "error", "reason": "等效分记录.xlsx 不存在"}

    try:
        wb = load_workbook(path)
    except Exception as e:
        return {"status": "error", "reason": f"等效分记录.xlsx 读取失败: {e}"}

    if "等效分记录" not in wb.sheetnames:
        wb.close()
        return {"status": "error", "reason": "等效分记录.xlsx 中缺少「等效分记录」Sheet"}

    ws = wb["等效分记录"]

    # 重复录入检测：同一考试名+日期不重复写入
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2:
            continue
        existing_name = str(row[0]).strip() if row[0] else ""
        existing_date = str(row[1]).strip() if row[1] else ""
        if existing_name == str(exam_name).strip() and existing_date == str(exam_date).strip():
            wb.close()
            return {
                "status": "error",
                "reason": f"已存在相同等效分记录（{exam_name} / {exam_date}），请勿重复保存",
            }

    # Build cross-validation columns
    cv_method1 = ""
    cv_score1 = ""
    cv_method2 = ""
    cv_score2 = ""
    cross = calc_result.get("cross_validations") or []
    if len(cross) > 0:
        cv_method1 = cross[0].get("method", "")
        cv_score1 = cross[0].get("score", "")
    if len(cross) > 1:
        cv_method2 = cross[1].get("method", "")
        cv_score2 = cross[1].get("score", "")

    gap = None
    eq_score = calc_result.get("equivalent_score")
    if target_line is not None and eq_score is not None:
        try:
            gap = round(float(eq_score) - float(target_line), 1)
        except (ValueError, TypeError):
            gap = None

    # I5: 去掉置信度中的"级"后缀，存储纯 A/B/C/D
    confidence = str(calc_result.get("confidence") or "").replace("级", "")

    # I6: 统一 calculation_detail 为字符串
    calc_detail = calc_result.get("calculation_detail", "")
    if isinstance(calc_detail, list):
        calc_detail = "|".join(str(x) for x in calc_detail)
    elif not isinstance(calc_detail, str):
        calc_detail = str(calc_detail)

    # I7: 统一 subject_scores 为 list-of-dict
    subject_scores = calc_result.get("subject_scores", [])
    if isinstance(subject_scores, dict):
        subject_scores = [{"subject": k, "score": v} for k, v in subject_scores.items()]
    elif not isinstance(subject_scores, list):
        subject_scores = []

    extra_info = json.dumps({
        "subject_scores": subject_scores,
        "warnings": calc_result.get("warnings") or [],
        "trust_note": calc_result.get("trust_note"),
        "divergence": calc_result.get("divergence"),
        "calculation_detail": calc_detail,
        "method_details": calc_result.get("method_details") or [],
    }, ensure_ascii=False)

    saved_row = None
    try:
        ws.append([
            exam_name,
            exam_date,
            calc_result.get("equivalent_score") if calc_result.get("equivalent_score") is not None else "",
            confidence,
            calc_result.get("primary_method") or "",
            cv_method1, cv_score1,
            cv_method2, cv_score2,
            calc_result.get("error_lower") if calc_result.get("error_lower") is not None else "",
            calc_result.get("error_upper") if calc_result.get("error_upper") is not None else "",
            target_university or "",
            target_line if target_line is not None else "",
            gap if gap is not None else "",
            extra_info,
        ])
        saved_row = ws.max_row
        wb.save(path)
    except Exception as e:
        wb.close()
        return {"status": "error", "reason": f"等效分记录写入失败: {e}"}
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return {
        "status": "ok",
        "row": saved_row,
        "score": calc_result.get("equivalent_score"),
        "confidence": calc_result.get("confidence"),
        "method": calc_result.get("primary_method"),
    }


def main():
    parser = argparse.ArgumentParser(description="Save equivalent score to Excel")
    parser.add_argument("--workspace", required=True, help="Workspace root path")
    parser.add_argument("--exam-name", required=True)
    parser.add_argument("--exam-date", required=True)
    parser.add_argument("--target", default=None, help="Target university name")
    parser.add_argument("--target-line", default=None, type=float, help="Target university admission score")
    args = parser.parse_args()

    try:
        calc_result = json.loads(sys.stdin.read())
    except json.JSONDecodeError as e:
        print(json.dumps({"status": "error", "reason": f"JSON 解析失败（请确认通过管道连接 calc_equivalent.py）: {e}"}, ensure_ascii=False))
        sys.exit(1)

    if not isinstance(calc_result, dict):
        print(json.dumps({"status": "error", "reason": "calc_equivalent 输出格式异常（非JSON对象）"}, ensure_ascii=False))
        sys.exit(1)

    if calc_result.get("status") not in ("ok",):
        print(json.dumps({"status": "skipped", "reason": calc_result.get("reason", "unknown")}, ensure_ascii=False))
        sys.exit(0)

    try:
        result = run(
            os.path.abspath(args.workspace),
            args.exam_name, args.exam_date,
            calc_result,
            args.target,
            args.target_line,
        )
    except Exception as e:
        result = {"status": "error", "reason": f"未知错误: {e}"}

    # 保存 max_row 到局部变量，避免 wb.close() 后访问
    if "row" in result:
        pass  # run() 已在 wb.close() 前获取
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
