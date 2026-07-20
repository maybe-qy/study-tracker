"""Integration test using real test data from 测试数据/ directory.

Verifies the full pipeline: setup → record → calc → generate,
and checks equivalent scores against manually calculated values.
"""

import json
import os
import sys
import tempfile
import shutil

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))
from setup_workspace import run as setup_workspace
from record_exam import run as record_exam
from calc_equivalent import run as calc_equivalent
from generate_reports import run as generate_reports


REAL_DATA = os.path.join(os.path.dirname(__file__), "..", "..", "测试数据")


def copy_real_macro_data(workspace):
    """Copy and reformat real macro data into workspace format."""
    from openpyxl import Workbook, load_workbook

    src_wb = load_workbook(os.path.join(REAL_DATA, "宏观数据.xlsx"), data_only=True)

    # ── 一分一段表 ──
    src_ws = src_wb["一分一段表_2026浙江"]
    score_rows = []
    for row in src_ws.iter_rows(min_row=3, values_only=True):  # skip title + headers
        score_val = row[0]
        cum_count = row[2]
        if score_val is not None and cum_count is not None:
            score_rows.append([float(score_val), int(cum_count), "浙江", 2026])

    # ── 特控线 ──
    src_ws = src_wb["特控线_浙江"]
    special_lines = []
    for row in src_ws.iter_rows(min_row=3, values_only=True):
        year, sl, _, _ = row
        if year is not None and sl is not None:
            special_lines.append([int(year), "浙江", float(sl)])

    # ── 赋分区间 ──
    src_ws = src_wb["赋分区间_浙江"]
    assign_ranges = []
    for row in src_ws.iter_rows(min_row=3, values_only=True):
        level, pct_range, score_range = row
        if level is not None and score_range is not None:
            parts = str(score_range).replace("~", "～").split("～")
            if len(parts) == 2:
                try:
                    lo, hi = int(parts[0]), int(parts[1])
                    assign_ranges.append(["浙江", int(level), lo, hi])
                except ValueError:
                    pass

    # ── 本校对照表（总分）── Use score distribution to create school lookup
    school_total_rows = []
    if score_rows:
        # Estimate: top student ≈ top score, each step down in rank ≈ score drop
        sorted_scores = sorted(score_rows, key=lambda r: r[1])  # by cum count
        for rank_pct in [1, 10, 50, 100, 200, 300, 500, 800, 1000, 1500, 2000, 3000, 5000]:
            target = rank_pct
            best_score = None
            best_diff = float("inf")
            for r in sorted_scores:
                diff = abs(r[1] - target)
                if diff < best_diff:
                    best_diff = diff
                    best_score = r[0]
            if best_score:
                school_total_rows.append([rank_pct, float(best_score)])

    # ── Write workspace macro file ──
    dst_path = os.path.join(workspace, "data", "macro", "宏观数据_只读.xlsx")
    wb = Workbook()

    def write_sheet(wb, name, headers, rows, first=False):
        if first:
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(title=name)
        ws.append(headers)
        for r in rows:
            ws.append(r)

    write_sheet(wb, "一分一段表", ["分数", "累计人数", "省份", "年份"], score_rows, first=True)
    write_sheet(wb, "特控线", ["年份", "省份", "特控线分数"], special_lines)
    write_sheet(wb, "赋分区间", ["省份", "等级", "最低分", "最高分"], assign_ranges)
    write_sheet(wb, "本校对照表_总分", ["校内排名", "高考总分"], school_total_rows)
    wb.save(dst_path)

    # Also create school admission data
    school_path = os.path.join(workspace, "data", "school", "学校招生_只读.xlsx")
    wb2 = Workbook()
    ws = wb2.active
    ws.title = "深大AI录取数据"
    ws.append(["年份", "专业", "录取最低分", "录取最低位次"])
    ws.append([2025, "人工智能", 652, "待查"])
    wb2.save(school_path)


def test_setup_workspace():
    """Test workspace setup with real data."""
    tmpdir = tempfile.mkdtemp()
    try:
        result = setup_workspace(tmpdir)
        assert result["status"] == "ok"
        # Verify all expected directories exist
        for sub in ["data/macro", "data/personal", "data/school"]:
            assert os.path.isdir(os.path.join(tmpdir, sub))
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def test_record_exam_9plus1():
    """Record the 9+1 midterm exam from real data."""
    tmpdir = tempfile.mkdtemp()
    try:
        setup_workspace(tmpdir)

        data = {
            "workspace": tmpdir,
            "exam_name": "2026.4高一下9+1期中考试",
            "exam_date": "2026-04",
            "exam_type": "联盟考试",
            "grade": "高一",
            "total_score": 589,
            "cn_score": 102.5,
            "math_score": 105,
            "en_score": 112.5,
            "sub1_name": "物理",
            "sub1_raw": 63,
            "sub1_assigned": 85,
            "sub1_confidence": "A",
            "sub2_name": "化学",
            "sub2_raw": 73,
            "sub2_assigned": 88,
            "sub2_confidence": "A",
            "sub3_name": "技术",
            "sub3_raw": 90,
            "sub3_assigned": 96,
            "sub3_confidence": "A",
            "special_line": 542.5,
            "notes": "技术赋分96经9+1联盟统一赋分(A级)",
        }

        result = record_exam(data)
        assert result["status"] == "ok"
        assert result["row"] == 2

        # Verify MD was created
        assert os.path.exists(result["md_path"])

        # Verify Excel data
        from openpyxl import load_workbook
        wb = load_workbook(os.path.join(tmpdir, "data", "personal", "成绩总表.xlsx"))
        ws = wb["成绩总表"]
        assert ws.cell(2, 1).value == "2026.4高一下9+1期中考试"
        assert ws.cell(2, 20).value == 589  # total score
        assert ws.cell(2, 25).value == 542.5  # 特控线
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def test_record_exam_final():
    """Record the final exam from real data (only 语数英 scores, 450-point scale)."""
    tmpdir = tempfile.mkdtemp()
    try:
        setup_workspace(tmpdir)

        data = {
            "workspace": tmpdir,
            "exam_name": "高一下期末考试",
            "exam_date": "2026-06",
            "exam_type": "全市统考",
            "grade": "高一",
            "total_score": 314,  # 语数英 only, 450-point scale
            "cn_score": 98.5,
            "math_score": 111,
            "en_score": 104.5,
            "sub1_name": "物理",
            "sub1_raw": 79.5,
            "sub1_assigned": 89,
            "sub1_confidence": "C",
            "sub2_name": "化学",
            "sub2_raw": 74,
            "sub2_assigned": 93,
            "sub2_confidence": "C",
            "sub3_name": "技术",
            "sub3_raw": 66,
            "sub3_confidence": "D",
            "special_line": 286.5,  # 450-point scale special line
            "notes": "选科无独立划线;C级置信度;总分=语数英314(450制)",
        }

        result = record_exam(data)
        assert result["status"] == "ok"
        assert result["row"] == 2
        assert os.path.exists(result["md_path"])
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def test_calc_equivalent_9plus1():
    """Calculate equivalent score for 9+1 exam. Expected: ~627."""
    tmpdir = tempfile.mkdtemp()
    try:
        setup_workspace(tmpdir)
        copy_real_macro_data(tmpdir)

        data = {
            "workspace": tmpdir,
            "total_score": 589,
            "special_line_exam": 542.5,
        }

        result = calc_equivalent(data)
        assert result["status"] == "ok"
        assert result["primary_method"] == "分数线对照法"
        assert result["confidence"] == "A"

        # Expected: ~627 (from manual calculation in MD)
        score = result["equivalent_score"]
        assert 620 <= score <= 635, f"Expected ~627, got {score}"

        print(f"\n  9+1期中: 等效分={score}, 方法={result['primary_method']}, 置信度={result['confidence']}")
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def test_calc_equivalent_final():
    """Calculate equivalent score for final exam. Expected: ~619."""
    tmpdir = tempfile.mkdtemp()
    try:
        setup_workspace(tmpdir)
        copy_real_macro_data(tmpdir)

        # Convert 450-scale to 750-scale for the calculation
        total_450 = 314
        sl_450 = 286.5
        total_750 = total_450 * 750 / 450
        sl_750 = sl_450 * 750 / 450

        data = {
            "workspace": tmpdir,
            "total_score": total_750,
            "special_line_exam": sl_750,
        }

        result = calc_equivalent(data)
        assert result["status"] == "ok"
        assert result["primary_method"] == "分数线对照法"
        assert result["confidence"] == "A"

        score = result["equivalent_score"]
        # Expected: ~619 (from manual calculation in MD)
        assert 610 <= score <= 630, f"Expected ~619, got {score}"

        print(f"\n  期末: 等效分={score}, 方法={result['primary_method']}, 置信度={result['confidence']}")
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def test_full_pipeline():
    """End-to-end: setup → record both exams → generate reports."""
    tmpdir = tempfile.mkdtemp()
    try:
        # 1. Setup
        setup_workspace(tmpdir)
        copy_real_macro_data(tmpdir)

        # 2. Record 9+1 exam
        r1 = record_exam({
            "workspace": tmpdir,
            "exam_name": "2026.4高一下9+1期中考试",
            "exam_date": "2026-04",
            "exam_type": "联盟考试",
            "grade": "高一",
            "total_score": 589,
            "cn_score": 102.5, "math_score": 105, "en_score": 112.5,
            "sub1_name": "物理", "sub1_raw": 63, "sub1_assigned": 85, "sub1_confidence": "A",
            "sub2_name": "化学", "sub2_raw": 73, "sub2_assigned": 88, "sub2_confidence": "A",
            "sub3_name": "技术", "sub3_raw": 90, "sub3_assigned": 96, "sub3_confidence": "A",
            "special_line": 542.5,
        })
        assert r1["status"] == "ok"

        # 3. Record final exam
        r2 = record_exam({
            "workspace": tmpdir,
            "exam_name": "高一下期末考试",
            "exam_date": "2026-06",
            "exam_type": "全市统考",
            "grade": "高一",
            "total_score": 314,
            "cn_score": 98.5, "math_score": 111, "en_score": 104.5,
            "sub1_name": "物理", "sub1_raw": 79.5, "sub1_assigned": 89, "sub1_confidence": "C",
            "sub2_name": "化学", "sub2_raw": 74, "sub2_assigned": 93, "sub2_confidence": "C",
            "sub3_name": "技术", "sub3_raw": 66, "sub3_confidence": "D",
            "special_line": 286.5,
        })
        assert r2["status"] == "ok"

        # 4. Generate reports (even with limited data, should not crash)
        gen_result = generate_reports(tmpdir)
        assert gen_result["status"] == "ok"
        assert len(gen_result["files"]) == 8

        # Verify all files exist and have content
        for f in gen_result["files"]:
            assert os.path.exists(f), f"Missing: {f}"
            size = os.path.getsize(f)
            assert size > 100, f"File too small: {f} ({size} bytes)"

        print(f"\n  生成报告: {len(gen_result['files'])} 个文件")
        for f in gen_result["files"]:
            print(f"    {os.path.basename(f)} ({os.path.getsize(f)} bytes)")

    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def test_cross_validation_both_methods():
    """Test that when both special_line and ranking are available, cross-validation works."""
    tmpdir = tempfile.mkdtemp()
    try:
        setup_workspace(tmpdir)
        copy_real_macro_data(tmpdir)

        data = {
            "workspace": tmpdir,
            "total_score": 589,
            "special_line_exam": 542.5,  # enables method 1
            "alliance_rank": 8000,       # enables method 3
            "alliance_total": 50000,
        }

        result = calc_equivalent(data)
        assert result["status"] == "ok"
        assert result["primary_method"] == "排名锚定法"
        assert len(result["cross_validations"]) >= 1

        print(f"\n  交叉验证: 主方法={result['primary_method']}, "
              f"交叉验证={[c['method'] for c in result['cross_validations']]}")
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)
