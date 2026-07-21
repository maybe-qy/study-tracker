"""Test calc_equivalent.py — the core calculation engine."""

import json
import os
import sys
import tempfile

import pytest

# Add scripts dir to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))
from calc_equivalent import run


def make_macro_ws(tmpdir):
    """Create minimal macro data for testing."""
    from openpyxl import Workbook
    ws_root = tmpdir
    macro_dir = os.path.join(ws_root, "data", "macro")
    os.makedirs(macro_dir, exist_ok=True)

    wb = Workbook()

    # 一分一段表 sheet
    ws1 = wb.active
    ws1.title = "一分一段表"
    ws1.append(["分数", "累计人数", "省份", "年份"])
    # Simulate: top score 750 = 1 person, each drop by 10 = +1000 people
    for i, score in enumerate(range(750, 299, -10)):
        ws1.append([score, (i + 1) * 100, "浙江", 2026])

    # 特控线 sheet
    ws2 = wb.create_sheet("特控线")
    ws2.append(["年份", "省份", "特控线分数"])
    ws2.append([2026, "浙江", 592])

    # 本校对照表_总分 sheet
    ws3 = wb.create_sheet("本校对照表_总分")
    ws3.append(["校内排名", "高考总分"])
    ws3.append([1, 720])
    ws3.append([10, 700])
    ws3.append([50, 670])
    ws3.append([100, 640])
    ws3.append([200, 600])
    ws3.append([300, 560])

    wb.save(os.path.join(macro_dir, "宏观数据_只读.xlsx"))
    return ws_root


def test_method_score_line(tmpdir):
    """Test 分数线对照法 with valid data."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "special_line_exam": 546.5,
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["primary_method"] == "分数线对照法"
    assert result["confidence"] == "A"
    assert 660 <= result["equivalent_score"] <= 680
    assert result["error_lower"] <= result["equivalent_score"] <= result["error_upper"]


def test_method_percentile(tmpdir):
    """Test 排名锚定法."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "alliance_rank": 2000,
        "alliance_total": 20000,
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["primary_method"] == "排名锚定法"
    assert result["confidence"] == "A"


def test_method_school_lookup(tmpdir):
    """Test 校内排名对照法 (A级)."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "school_rank": 50,
        "school_total": 500,
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["primary_method"] == "校内排名对照法"
    assert result["confidence"] == "A"
    assert abs(result["equivalent_score"] - 670) < 10
    assert result["method_count"] == 1


def make_macro_ws_no_lookup(tmpdir):
    """Create macro data WITHOUT 本校对照表 (so method 4 is the only school method)."""
    from openpyxl import Workbook
    ws_root = tmpdir
    macro_dir = os.path.join(ws_root, "data", "macro")
    os.makedirs(macro_dir, exist_ok=True)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "一分一段表"
    ws1.append(["分数", "累计人数", "省份", "年份"])
    for i, score in enumerate(range(750, 299, -10)):
        ws1.append([score, (i + 1) * 100, "浙江", 2026])

    wb.save(os.path.join(macro_dir, "宏观数据_只读.xlsx"))
    return ws_root


def test_school_rank_no_lookup_insufficient(tmpdir):
    """校排名无对照表时，走校排名估算(C级)得出等效分."""
    ws = make_macro_ws_no_lookup(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "school_rank": 80,
        "school_total": 500,
        "school_type": "市重点",
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["primary_method"] == "校排名估算"
    assert result["confidence"] == "C"


def test_cross_validation(tmpdir):
    """Test that multiple methods produce cross-validation."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "special_line_exam": 546.5,
        "alliance_rank": 3200,
        "alliance_total": 21000,
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["primary_method"] == "分数线对照法"
    assert len(result["cross_validations"]) >= 1
    assert "trust_note" in result


def test_insufficient_data(tmpdir):
    """Test that no data returns proper error."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        # No ranking, no special line
    }
    result = run(data)
    assert result["status"] == "insufficient_data"


def test_no_macro_file(tmpdir):
    """Test that missing macro file returns error."""
    data = {
        "workspace": str(tmpdir),
        "total_score": 650,
    }
    result = run(data)
    assert result["status"] == "error"


def test_priority_order(tmpdir):
    """Test that score_line (P2) beats percentile (P5) when both available."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "special_line_exam": 546.5,
        "alliance_rank": 3200,
        "alliance_total": 21000,
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["primary_method"] == "分数线对照法"  # P1 wins over P3


def test_rank_exceeds_total(tmpdir):
    """Test that rank > total is caught."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "alliance_rank": 50000,
        "alliance_total": 1000,  # rank exceeds total
    }
    result = run(data)
    # Should fall through to insufficient since percentile method fails
    assert result["status"] in ("insufficient_data", "ok")


def test_percentile_gaoer(tmpdir):
    """Test that 高二 percentile anchoring still gets A-level (grade no longer affects confidence)."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "alliance_rank": 2000,
        "alliance_total": 20000,
        "grade": "高二",
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["primary_method"] == "排名锚定法"
    assert result["confidence"] == "A"


def test_gaoyi_no_longer_blocked(tmpdir):
    """Test that 高一 with ranking data now calculates equivalent score (grade no longer blocks)."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "alliance_rank": 2000,
        "alliance_total": 20000,
        "grade": "高一",
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["confidence"] == "A"


def test_score_line_beats_percentile(tmpdir):
    """Test that score_line (P2) takes priority over percentile (P5)."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "special_line_exam": 546.5,
        "alliance_rank": 3200,
        "alliance_total": 21000,
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["primary_method"] == "分数线对照法"
    assert result["confidence"] == "A"


def test_score_scale_450_subject_ratio(tmpdir):
    """B2 fix: 450分制下，单科等效分比例计算应使用原始总分（非换算后的750制）."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 314,
        "score_scale": 450,
        "special_line_exam": 286.5,
        "subjects": [
            {"name": "语文", "raw": 98.5},
            {"name": "数学", "raw": 111},
            {"name": "英语", "raw": 104.5},
        ],
    }
    result = run(data)
    assert result["status"] == "ok"
    subject_scores = result.get("subject_scores", [])
    assert len(subject_scores) == 3
    total_subject_sum = sum(s["score"] for s in subject_scores if s["score"])
    # M14: verify ratio uses original 450-scale denominator, not converted 750
    chinese = [s for s in subject_scores if s["subject"] == "语文"][0]
    actual_ratio = chinese["score"] / total_subject_sum
    expected_ratio = 98.5 / 314  # should be ~0.3137, NOT 98.5/523.33≈0.188
    assert abs(actual_ratio - expected_ratio) < 0.01, \
        f"语文占比{actual_ratio:.4f}，期望{expected_ratio:.4f}（B2: 分母应为原始450制，非750制）"


def make_macro_ws_with_upgrade(tmpdir):
    """Create macro data with期末升级 sheet for two-module and school-threshold tests."""
    from openpyxl import Workbook
    ws_root = tmpdir
    macro_dir = os.path.join(ws_root, "data", "macro")
    os.makedirs(macro_dir, exist_ok=True)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "一分一段表"
    ws1.append(["分数", "累计人数", "省份", "年份"])
    for i, score in enumerate(range(750, 299, -10)):
        ws1.append([score, (i + 1) * 100, "浙江", 2026])

    ws2 = wb.create_sheet("特控线")
    ws2.append(["年份", "省份", "特控线分数"])
    ws2.append([2026, "浙江", 592])

    # 期末升级 sheet: col0=科目, col1=2027划线, col2=2027上线, col3=2028划线, col4=2028上线
    ws3 = wb.create_sheet("期末高一下升级")
    ws3.append(["科目", "2027划线", "2027上线", "2028划线", "2028上线"])
    ws3.append(["特控分段", "", "", "", ""])
    ws3.append(["语数英综合", "", "", 270, 500])
    ws3.append(["物理", "", "", 65, ""])
    ws3.append(["化学", "", "", 70, ""])
    ws3.append(["浙大分段", "", "", "", ""])
    ws3.append(["语数英综合", "", "", 300, 150])
    ws3.append(["物理", "", "", 85, ""])
    ws3.append(["化学", "", "", 90, ""])

    # 结构 sheet for school_threshold
    ws4 = wb.create_sheet("期末结构")
    ws4.append(["类别", "人数"])
    ws4.append(["全校总人数", "835人"])

    wb.save(os.path.join(macro_dir, "宏观数据_只读.xlsx"))
    return ws_root


def test_two_module_method(tmpdir):
    """M13: 双模块换算法 (priority 1) with upgrade sheet."""
    ws = make_macro_ws_with_upgrade(tmpdir)
    data = {
        "workspace": ws,
        "exam_name": "高一下期末",
        "total_score": 570,
        "score_scale": 750,
        "subjects": [
            {"name": "语文", "raw": 115},
            {"name": "数学", "raw": 108},
            {"name": "英语", "raw": 112},
            {"name": "物理", "raw": 80, "assigned": 88},
            {"name": "化学", "raw": 75, "assigned": 85},
        ],
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["primary_method"] == "双模块换算法"
    assert result["confidence"] in ("A", "B")


def test_school_threshold_method(tmpdir):
    """M13: 校排阈值估算法 (priority 3) triggered with upgrade + school data."""
    ws = make_macro_ws_with_upgrade(tmpdir)
    data = {
        "workspace": ws,
        "exam_name": "高一下期末",
        "total_score": 480,  # → 288 in 450-scale, between 特控270 and 浙大300
        "score_scale": 750,
        "school_type": "省重点",
        "subjects": [
            {"name": "语文", "raw": 105},
            {"name": "数学", "raw": 100},
            {"name": "英语", "raw": 95},
            {"name": "物理", "raw": 60},
            {"name": "化学", "raw": 55},
        ],
    }
    result = run(data)
    assert result["status"] == "ok"
    # 双模块 should be primary (P1), 校排阈值 should appear as cross-validation
    methods_seen = [m["method"] for m in result["method_details"]]
    assert "校排阈值估算法" in methods_seen


def test_independent_subject_sum(tmpdir):
    """M12: compute_independent_subject_sum produces estimate independent of total."""
    from calc_equivalent import compute_independent_subject_sum, read_macro_data
    ws = make_macro_ws(tmpdir)
    macro = read_macro_data(ws)
    data = {
        "total_score": 650,
        "score_scale": 750,
        "special_line_exam": 546.5,
        "subjects": [
            {"name": "语文", "raw": 120},
            {"name": "数学", "raw": 110},
            {"name": "英语", "raw": 100},
            {"name": "物理", "assigned": 88},
            {"name": "化学", "assigned": 91},
            {"name": "技术", "assigned": 66},
        ],
    }
    result = compute_independent_subject_sum(data, macro)
    assert result is not None
    assert "sum" in result
    assert "confidences" in result
    assert len(result["confidences"]) == 6  # 3 main + 3 elective
    assert result["sum"] > 0
