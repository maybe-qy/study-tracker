"""Test record_exam.py — exam recording and validation."""

import json
import os
import sys
import tempfile

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "scripts"))
from record_exam import validate, build_row, MD_TEMPLATE
from setup_workspace import run as setup_workspace


def make_workspace(tmpdir):
    """Create a minimal workspace with Excel files."""
    ws = str(tmpdir)
    setup_workspace(ws)
    return ws


def test_validate_missing_required():
    """Test that missing required fields are caught."""
    data = {"exam_name": "test"}
    err = validate(data)
    assert err is not None
    assert "workspace" in err or "exam_date" in err


def test_validate_all_required():
    """Test that all required fields pass validation."""
    data = {
        "workspace": ".",
        "exam_name": "期末",
        "exam_date": "2026-01",
        "exam_type": "全市统考",
        "grade": "高一",
        "total_score": 650,
    }
    err = validate(data)
    assert err is None


def test_build_row():
    """Test row mapping from JSON to Excel column order."""
    data = {
        "exam_name": "期末",
        "exam_date": "2026-01",
        "exam_type": "全市统考",
        "grade": "高一",
        "total_score": 650,
        "cn_score": 118,
        "math_score": 135,
        "en_score": 128,
        "sub1_name": "物理",
        "sub1_raw": 78,
        "sub1_assigned": 91,
        "sub1_confidence": "A",
        "alliance_rank": 3200,
        "alliance_total": 21000,
    }
    row = build_row(data)
    assert row[0] == "期末"
    assert row[4] == 118  # 语文
    assert row[5] == 135  # 数学
    assert row[6] == 128  # 英语
    assert row[19] == 650    # 总分
    assert row[20] == 3200   # 市/联盟排名 (from alliance_rank)
    assert row[21] == 21000  # 市/联盟总人数


def test_record_exam_end_to_end(tmpdir):
    """End-to-end: setup workspace -> record exam -> verify files exist."""
    from record_exam import run as record_exam

    ws = make_workspace(tmpdir)
    data = {
        "workspace": ws,
        "exam_name": "高一期末",
        "exam_date": "2026-01",
        "exam_type": "全市统考",
        "grade": "高一",
        "total_score": 650,
        "cn_score": 118,
        "math_score": 135,
        "en_score": 128,
        "sub1_name": "物理",
        "sub1_raw": 78,
        "sub1_assigned": 91,
        "sub1_confidence": "A",
    }
    result = record_exam(data)
    assert result["status"] == "ok"
    assert result["row"] == 2  # First data row after headers

    # Verify .md file was created
    assert os.path.exists(result["md_path"])

    # Verify Excel has the data
    from openpyxl import load_workbook
    wb = load_workbook(os.path.join(ws, "data", "personal", "成绩总表.xlsx"))
    ws2 = wb["成绩总表"]
    assert ws2.max_row == 2
    assert ws2.cell(2, 1).value == "高一期末"


def test_total_sum_mismatch_marker(tmpdir):
    """Test that total vs sum discrepancy is marked in .md."""
    from record_exam import run as record_exam

    ws = make_workspace(tmpdir)
    data = {
        "workspace": ws,
        "exam_name": "期末",
        "exam_date": "2026-01",
        "exam_type": "校级考试",
        "grade": "高一",
        "total_score": 650,  # Total is 650
        "cn_score": 100,     # Sum = 100+100+100+80+80+80 = 540 ≠ 650
        "math_score": 100,
        "en_score": 100,
        "sub1_raw": 80,
        "sub2_raw": 80,
        "sub3_raw": 80,
    }
    result = record_exam(data)
    assert result["status"] == "ok"

    # Read .md and check for the discrepancy note
    with open(result["md_path"], "r", encoding="utf-8") as f:
        content = f.read()
    assert "用户确认以总分为准" in content
