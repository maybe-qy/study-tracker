"""Test new features: duplicate detection, exception handling, save_equivalent."""

import json
import os
import sys

import pytest
from openpyxl import load_workbook

from setup_workspace import run as setup_workspace
from record_exam import run as record_exam, _check_duplicate
from save_equivalent import run as save_equivalent


def make_workspace(tmpdir):
    """Create a minimal workspace with Excel files."""
    ws = str(tmpdir)
    setup_workspace(ws)
    return ws


def _sample_exam_data(ws):
    return {
        "workspace": ws,
        "exam_name": "11月期中",
        "exam_date": "2025-11",
        "exam_type": "期中",
        "grade": "高二",
        "total_score": 576,
        "cn_score": 102,
        "math_score": 128,
        "en_score": 110,
        "sub1_name": "物理",
        "sub1_raw": 70,
        "sub1_assigned": 82,
        "sub1_confidence": "B",
        "sub2_name": "化学",
        "sub2_raw": 68,
        "sub2_assigned": 79,
        "sub2_confidence": "B",
        "sub3_name": "技术",
        "sub3_raw": 88,
        "sub3_assigned": 84,
        "sub3_confidence": "B",
        "school_rank": 150,
        "school_total": 600,
    }


# ── Duplicate detection ──────────────────────────────────────────────


def test_duplicate_detection(tmpdir):
    """Recording the same exam twice should be rejected."""
    ws = make_workspace(tmpdir)
    data = _sample_exam_data(ws)

    # First recording succeeds
    result1 = record_exam(data)
    assert result1["status"] == "ok"
    assert result1["record_index"] == 1

    # Second recording of same exam+date should fail
    result2 = record_exam(data)
    assert result2["status"] == "error"
    assert "重复" in result2["reason"] or "已存在" in result2["reason"]


def test_different_exams_allowed(tmpdir):
    """Recording different exams should both succeed."""
    ws = make_workspace(tmpdir)

    data1 = _sample_exam_data(ws)
    result1 = record_exam(data1)
    assert result1["status"] == "ok"

    data2 = _sample_exam_data(ws)
    data2["exam_name"] = "12月月考"
    data2["exam_date"] = "2025-12"
    result2 = record_exam(data2)
    assert result2["status"] == "ok"
    assert result2["record_index"] == 2


def test_check_duplicate_directly(tmpdir):
    """Test _check_duplicate helper directly."""
    ws = make_workspace(tmpdir)
    excel_path = os.path.join(ws, "data", "personal", "成绩总表.xlsx")
    wb = load_workbook(excel_path)
    ws_sheet = wb["成绩总表"]

    # Empty sheet — no duplicate
    assert _check_duplicate(ws_sheet, "任何考试", "2025-11") is False

    # Add a row
    ws_sheet.append(["11月期中", "2025-11", "期中", "高二"])
    wb.save(excel_path)
    wb.close()

    wb = load_workbook(excel_path)
    ws_sheet = wb["成绩总表"]
    assert _check_duplicate(ws_sheet, "11月期中", "2025-11") is True
    assert _check_duplicate(ws_sheet, "11月期中", "2025-12") is False
    assert _check_duplicate(ws_sheet, "12月月考", "2025-11") is False
    wb.close()


# ── Exception handling ───────────────────────────────────────────────


def test_missing_sheet_error(tmpdir):
    """Recording into a workspace with missing Sheet should return error."""
    ws = make_workspace(tmpdir)
    excel_path = os.path.join(ws, "data", "personal", "成绩总表.xlsx")

    # Corrupt the Excel by renaming the sheet (can't delete only sheet)
    wb = load_workbook(excel_path)
    wb["成绩总表"].title = "wrong_sheet"
    wb.save(excel_path)
    wb.close()

    data = _sample_exam_data(ws)
    result = record_exam(data)
    assert result["status"] == "error"
    assert "成绩总表" in result["reason"]


def test_missing_excel_error(tmpdir):
    """Recording into a workspace without Excel should return clean error."""
    ws = make_workspace(tmpdir)
    os.remove(os.path.join(ws, "data", "personal", "成绩总表.xlsx"))

    data = _sample_exam_data(ws)
    result = record_exam(data)
    assert result["status"] == "error"
    assert "不存在" in result["reason"]


# ── save_equivalent tests ────────────────────────────────────────────


def test_save_equivalent_basic(tmpdir):
    """Save a basic equivalent score result."""
    ws = make_workspace(tmpdir)

    # First record an exam
    record_exam(_sample_exam_data(ws))

    # Save equivalent score
    calc_result = {
        "status": "ok",
        "equivalent_score": 580.5,
        "confidence": "B级",
        "primary_method": "分数线对照法",
        "error_lower": 570,
        "error_upper": 590,
        "cross_validations": [],
        "subject_scores": [],
        "warnings": [],
        "calculation_detail": "测试计算",
    }

    result = save_equivalent(
        ws, "11月期中", "2025-11", calc_result,
    )
    assert result["status"] == "ok"
    assert result["score"] == 580.5

    # Verify it was saved
    wb = load_workbook(os.path.join(ws, "data", "personal", "等效分记录.xlsx"))
    eq_ws = wb["等效分记录"]
    assert eq_ws.max_row == 2  # header + 1 data
    assert eq_ws.cell(2, 1).value == "11月期中"
    assert eq_ws.cell(2, 3).value == 580.5
    wb.close()


def test_save_equivalent_missing_file(tmpdir):
    """Save equivalent should return error if file missing."""
    ws = make_workspace(tmpdir)
    os.remove(os.path.join(ws, "data", "personal", "等效分记录.xlsx"))

    result = save_equivalent(ws, "test", "2025-01", {"status": "ok", "equivalent_score": 500})
    assert result["status"] == "error"
    assert "不存在" in result["reason"]


def test_save_equivalent_with_target(tmpdir):
    """Save equivalent with target university info."""
    ws = make_workspace(tmpdir)
    record_exam(_sample_exam_data(ws))

    calc_result = {
        "status": "ok",
        "equivalent_score": 600,
        "confidence": "A级",
        "primary_method": "双模块换算法",
        "error_lower": 595,
        "error_upper": 605,
        "cross_validations": [],
        "subject_scores": [],
        "warnings": [],
        "calculation_detail": "双模块",
    }

    result = save_equivalent(
        ws, "11月期中", "2025-11", calc_result,
        target_university="浙江大学",
        target_line=652,
    )
    assert result["status"] == "ok"

    # Verify target info was saved
    wb = load_workbook(os.path.join(ws, "data", "personal", "等效分记录.xlsx"))
    eq_ws = wb["等效分记录"]
    assert eq_ws.cell(2, 12).value == "浙江大学"   # col L = target_university
    assert eq_ws.cell(2, 14).value == -52          # col N = gap = 600 - 652
    wb.close()


# ── Windows filename safety ──────────────────────────────────────────


def test_safe_name_windows_chars(tmpdir):
    """Test that Windows-illegal characters in exam names are sanitized."""
    ws = make_workspace(tmpdir)

    data = _sample_exam_data(ws)
    data["exam_name"] = "11月:期中?"
    data["exam_date"] = "2025-11"

    result = record_exam(data)
    assert result["status"] == "ok"
    # MD file should exist with sanitized name
    assert os.path.exists(result["md_path"])
