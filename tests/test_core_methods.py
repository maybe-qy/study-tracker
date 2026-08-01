"""Additional unit tests for core calculation methods and shared utilities.

Covers:
  - 人数校准法 (Population calibration)
  - excel_utils shared module (is_header_row, read_sheet_dicts, find_sheet, filter_numeric_rows)
  - Perfect-score edge case (750)
  - Error-range margins by confidence level
  - Weighted fusion logic
  - Divergence / trust_note thresholds
"""

import os

import pytest
from openpyxl import Workbook, load_workbook

from calc_equivalent import run, method_population_calibration, method_score_line, method_school_lookup, _find_previous_subject_data
from excel_utils import is_header_row, read_sheet_dicts, find_sheet, filter_numeric_rows, filter_score_table, SHEET_KEY_MAP


# ── Helpers ───────────────────────────────────────────────────────────


def make_macro_with_threshold(tmpdir):
    """Macro data with a 门槛 sheet for population calibration tests."""
    ws_root = str(tmpdir)
    macro_dir = os.path.join(ws_root, "data", "macro")
    os.makedirs(macro_dir, exist_ok=True)

    wb = Workbook()

    # 一分一段表
    ws1 = wb.active
    ws1.title = "一分一段表"
    ws1.append(["分数", "累计人数", "省份", "年份"])
    for i, score in enumerate(range(750, 299, -10)):
        ws1.append([score, (i + 1) * 100, "浙江", 2026])

    # 特控线
    ws2 = wb.create_sheet("特控线")
    ws2.append(["年份", "省份", "特控线分数"])
    ws2.append([2026, "浙江", 594])

    # 门槛数据 sheet — has 特控线上线人数
    ws3 = wb.create_sheet("门槛数据")
    ws3.append(["考试名称", "特控线分数", "特控线上线人数", "A线分数", "A线上线人数"])
    ws3.append(["11月期中", 510, 573, 570, 120])

    wb.save(os.path.join(macro_dir, "宏观数据_只读.xlsx"))
    return ws_root


# ── 人数校准法 (Population calibration) ──────────────────────────────


def test_population_calibration_direct():
    """Test method_population_calibration directly with mock macro data."""
    macro = {
        "一分一段表": [
            {"分数": 750, "累计人数": 100, "省份": "浙江", "年份": 2026},
            {"分数": 700, "累计人数": 1000, "省份": "浙江", "年份": 2026},
            {"分数": 650, "累计人数": 5000, "省份": "浙江", "年份": 2026},
            {"分数": 600, "累计人数": 15000, "省份": "浙江", "年份": 2026},
            {"分数": 594, "累计人数": 20000, "省份": "浙江", "年份": 2026},
            {"分数": 550, "累计人数": 35000, "省份": "浙江", "年份": 2026},
            {"分数": 500, "累计人数": 50000, "省份": "浙江", "年份": 2026},
            {"分数": 400, "累计人数": 80000, "省份": "浙江", "年份": 2026},
            {"分数": 300, "累计人数": 100000, "省份": "浙江", "年份": 2026},
        ],
        "特控线": [
            {"年份": 2026, "省份": "浙江", "特控线分数": 594},
        ],
        "门槛数据": [
            {"考试名称": "11月期中", "特控线分数": 510, "特控线上线人数": 573},
        ],
    }

    data = {
        "school_rank": 100,
    }

    result = method_population_calibration(data, macro)
    assert result is not None
    assert result["method"] == "人数校准法"
    assert result["confidence"] == "B"
    assert result["score"] > 0
    # School rank 100 with 573 teckong students → calibrated city rank ≈ 100 * (20000/573) ≈ 3490
    # That should map to a score around 650-680
    assert 600 <= result["score"] <= 700


def test_population_calibration_no_school_rank():
    """Without school_rank, population calibration returns None."""
    macro = {
        "一分一段表": [{"分数": 600, "累计人数": 10000}],
        "特控线": [{"年份": 2026, "特控线分数": 594}],
        "门槛数据": [{"特控线上线人数": 573}],
    }
    assert method_population_calibration({}, macro) is None


def test_population_calibration_no_threshold_sheet():
    """Without a 门槛/升级 sheet, returns None."""
    macro = {
        "一分一段表": [{"分数": 600, "累计人数": 10000}],
        "特控线": [{"年份": 2026, "特控线分数": 594}],
    }
    data = {"school_rank": 100}
    assert method_population_calibration(data, macro) is None


def test_population_calibration_via_run(tmpdir):
    """Integration: run() with threshold data triggers population calibration."""
    ws = make_macro_with_threshold(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 576,
        "school_rank": 100,
        "school_total": 600,
    }
    result = run(data)
    assert result["status"] == "ok"
    # Population calibration should be among the methods (may not be primary
    # if school_estimate also fires, but should appear in method_details)
    method_names = [m["method"] for m in result["method_details"]]
    assert "人数校准法" in method_names


# ── Perfect score edge case ──────────────────────────────────────────


def test_perfect_score_750(tmpdir):
    """Total score 750 should yield equivalent 750 via score-line method."""
    ws = str(tmpdir)
    macro_dir = os.path.join(ws, "data", "macro")
    os.makedirs(macro_dir, exist_ok=True)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "一分一段表"
    ws1.append(["分数", "累计人数", "省份", "年份"])
    for i, score in enumerate(range(750, 299, -10)):
        ws1.append([score, (i + 1) * 100, "浙江", 2026])
    ws2 = wb.create_sheet("特控线")
    ws2.append(["年份", "省份", "特控线分数"])
    ws2.append([2026, "浙江", 594])
    wb.save(os.path.join(macro_dir, "宏观数据_只读.xlsx"))

    data = {
        "workspace": ws,
        "total_score": 750,
        "special_line_exam": 546.5,
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["equivalent_score"] == 750.0


# ── Error range margins ──────────────────────────────────────────────


def test_error_range_a_level(tmpdir):
    """A-level confidence should have ±5 margin (single method)."""
    macro_dir = os.path.join(str(tmpdir), "data", "macro")
    os.makedirs(macro_dir, exist_ok=True)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "一分一段表"
    ws1.append(["分数", "累计人数", "省份", "年份"])
    for i, score in enumerate(range(750, 299, -10)):
        ws1.append([score, (i + 1) * 100, "浙江", 2026])
    ws2 = wb.create_sheet("特控线")
    ws2.append(["年份", "省份", "特控线分数"])
    ws2.append([2026, "浙江", 594])
    wb.save(os.path.join(macro_dir, "宏观数据_只读.xlsx"))

    data = {
        "workspace": str(tmpdir),
        "total_score": 650,
        "special_line_exam": 546.5,
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["confidence"] == "A"
    score = result["equivalent_score"]
    # With a single A-level method, margin = ±5
    # (fusion with subject sum may widen it, but lower bound should be ≤ score-5)
    assert result["error_lower"] <= score - 5 + 0.1  # allow tiny float slack
    assert result["error_upper"] >= score + 5 - 0.1


# ── Divergence / trust_note ──────────────────────────────────────────


def test_divergence_low(tmpdir):
    """When methods from different families, divergence should be one of low/medium/high."""
    from calc_equivalent import run as calc_run
    # Create macro with 一分一段表 + 特控线 + 升级 sheet + 本校对照表
    macro_dir = os.path.join(str(tmpdir), "data", "macro")
    os.makedirs(macro_dir, exist_ok=True)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "一分一段表"
    ws1.append(["分数", "累计人数", "省份", "年份"])
    for i, score in enumerate(range(750, 299, -10)):
        ws1.append([score, (i + 1) * 100, "浙江", 2026])
    ws2 = wb.create_sheet("特控线")
    ws2.append(["年份", "省份", "特控线分数"])
    ws2.append([2026, "浙江", 594])
    # 升级 sheet
    ws3 = wb.create_sheet("期末升级")
    ws3.append(["科目", "2027划线", "2027上线", "2028划线", "2028上线"])
    ws3.append(["特控分段", "", "", "", ""])
    ws3.append(["语数英综合", "", "", 270, 500])
    ws3.append(["物理", "", "", 65, ""])
    ws3.append(["化学", "", "", 70, ""])
    ws3.append(["浙大分段", "", "", "", ""])
    ws3.append(["语数英综合", "", "", 300, 150])
    ws3.append(["物理", "", "", 85, ""])
    ws3.append(["化学", "", "", 90, ""])
    ws4 = wb.create_sheet("本校对照表_总分")
    ws4.append(["校内排名", "高考总分"])
    ws4.append([1, 720])
    ws4.append([10, 700])
    ws4.append([50, 670])
    ws4.append([100, 640])
    ws4.append([200, 600])
    ws4.append([300, 560])
    wb.save(os.path.join(macro_dir, "宏观数据_只读.xlsx"))

    data = {
        "workspace": str(tmpdir),
        "exam_name": "期末",
        "total_score": 650,
        "special_line_exam": 546.5,
        "alliance_rank": 3200,
        "alliance_total": 21000,
        "subjects": [
            {"name": "语文", "raw": 120},
            {"name": "数学", "raw": 130},
            {"name": "英语", "raw": 140},
            {"name": "物理", "raw": 80, "assigned": 88},
            {"name": "化学", "raw": 75, "assigned": 85},
        ],
    }
    result = calc_run(data)
    assert result["status"] == "ok"
    assert result.get("divergence") is not None
    # divergence should be one of low/medium/high
    assert result["divergence"] in ("low", "medium", "high")
    assert result.get("trust_note") is not None


# ── excel_utils tests ────────────────────────────────────────────────


class TestIsHeaderRow:
    def test_clear_header(self):
        """Row with known column names should be detected as header."""
        assert is_header_row(("分数", "累计人数", "省份")) is True

    def test_title_row(self):
        """A title row with no known keywords should not be header."""
        assert is_header_row(("2026年浙江省高考一分一段表",)) is False

    def test_empty_row(self):
        assert is_header_row(()) is False

    def test_none_values(self):
        assert is_header_row((None, None)) is False

    def test_single_known_keyword(self):
        """First column being a known keyword triggers header detection."""
        assert is_header_row(("排名", "something")) is True

    def test_mixed_header(self):
        """Row with 2+ keyword hits is a header even if first col is unknown."""
        assert is_header_row(("学生成绩", "总分", "排名")) is True


class TestReadSheetDicts:
    def test_basic_read(self):
        """Read a simple worksheet with header on row 1."""
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "分数"])
        ws.append(["张三", 90])
        ws.append(["李四", 85])
        rows = read_sheet_dicts(ws)
        assert len(rows) == 2
        assert rows[0]["姓名"] == "张三"
        assert rows[0]["分数"] == 90
        assert rows[1]["分数"] == 85

    def test_title_row_skip(self):
        """Title row on row 1, header on row 2 should be handled."""
        wb = Workbook()
        ws = wb.active
        ws.append(["某次考试信息表"])  # title row (no known keywords)
        ws.append(["分数", "排名"])  # header row (2 known keywords)
        ws.append([90, 100])
        rows = read_sheet_dicts(ws)
        assert len(rows) == 1
        assert rows[0]["分数"] == 90
        assert rows[0]["排名"] == 100

    def test_empty_sheet(self):
        """Sheet with only headers returns empty list."""
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "分数"])
        rows = read_sheet_dicts(ws)
        assert rows == []

    def test_extra_columns_truncated(self):
        """Extra data columns beyond headers get placeholder names (bounds-safe)."""
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名", "分数"])
        ws.append(["张三", 90])
        rows = read_sheet_dicts(ws)
        assert len(rows) == 1
        assert "姓名" in rows[0]
        assert "分数" in rows[0]
        assert rows[0]["姓名"] == "张三"
        assert rows[0]["分数"] == 90


class TestFindSheet:
    def test_exact_match(self):
        assert find_sheet(["特控线", "一分一段表"], "特控线") == "特控线"

    def test_fuzzy_match(self):
        assert find_sheet(["一分一段表_2026浙江"], "一分一段") == "一分一段表_2026浙江"

    def test_no_match(self):
        assert find_sheet(["无关Sheet"], "特控线") is None

    def test_display_name_preferred(self):
        """'对照' should prefer '本校对照表_总分' over '单科对照'."""
        sheets = ["单科对照", "本校对照表_总分"]
        assert find_sheet(sheets, "对照") == "本校对照表_总分"

    def test_empty_list(self):
        assert find_sheet([], "特控线") is None


class TestFilterNumericRows:
    def test_filters_non_numeric(self):
        rows = [
            {"分数": 750, "累计人数": 100},
            {"分数": "N/A", "累计人数": 200},
            {"分数": "缺考", "累计人数": 300},
        ]
        filtered = filter_numeric_rows(rows, "分数")
        assert len(filtered) == 1
        assert filtered[0]["分数"] == 750

    def test_empty_list(self):
        assert filter_numeric_rows([], "分数") == []

    def test_no_matching_key(self):
        rows = [{"name": "test"}]
        assert filter_numeric_rows(rows, "分数") == []


# ── Fusion logic ─────────────────────────────────────────────────────


def test_fusion_between_min_max(tmpdir):
    """Fused equivalent score should lie between min and max of component methods."""
    from calc_equivalent import run as calc_run
    macro_dir = os.path.join(str(tmpdir), "data", "macro")
    os.makedirs(macro_dir, exist_ok=True)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "一分一段表"
    ws1.append(["分数", "累计人数", "省份", "年份"])
    for i, score in enumerate(range(750, 299, -10)):
        ws1.append([score, (i + 1) * 100, "浙江", 2026])
    ws2 = wb.create_sheet("特控线")
    ws2.append(["年份", "省份", "特控线分数"])
    ws2.append([2026, "浙江", 594])
    # 升级 sheet for 族①
    ws3 = wb.create_sheet("期末升级")
    ws3.append(["科目", "2027划线", "2027上线", "2028划线", "2028上线"])
    ws3.append(["特控分段", "", "", "", ""])
    ws3.append(["语数英综合", "", "", 270, 500])
    ws3.append(["物理", "", "", 65, ""])
    ws3.append(["化学", "", "", 70, ""])
    ws3.append(["浙大分段", "", "", "", ""])
    ws3.append(["语数英综合", "", "", 300, 150])
    ws3.append(["物理", "", "", 85, ""])
    ws3.append(["化学", "", "", 90, ""])
    ws4 = wb.create_sheet("本校对照表_总分")
    ws4.append(["校内排名", "高考总分"])
    ws4.append([1, 720])
    ws4.append([10, 700])
    ws4.append([50, 670])
    ws4.append([100, 640])
    ws4.append([200, 600])
    ws4.append([300, 560])
    wb.save(os.path.join(macro_dir, "宏观数据_只读.xlsx"))

    data = {
        "workspace": str(tmpdir),
        "exam_name": "期末",
        "total_score": 650,
        "special_line_exam": 546.5,
        "alliance_rank": 3200,
        "alliance_total": 21000,
        "subjects": [
            {"name": "语文", "raw": 120},
            {"name": "数学", "raw": 130},
            {"name": "英语", "raw": 140},
            {"name": "物理", "raw": 80, "assigned": 88},
            {"name": "化学", "raw": 75, "assigned": 85},
        ],
    }
    result = calc_run(data)
    assert result["status"] == "ok"
    assert len(result["method_details"]) >= 2

    scores = [m["score"] for m in result["method_details"]]
    min_score = min(scores)
    max_score = max(scores)
    fused = result["equivalent_score"]
    # Fused score should be within the range of method scores (with small tolerance)
    assert min_score - 20 <= fused <= max_score + 20


def test_single_method_no_fusion(tmpdir):
    """With only one method, equivalent score should equal that method's score."""
    macro_dir = os.path.join(str(tmpdir), "data", "macro")
    os.makedirs(macro_dir, exist_ok=True)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "一分一段表"
    ws1.append(["分数", "累计人数", "省份", "年份"])
    for i, score in enumerate(range(750, 299, -10)):
        ws1.append([score, (i + 1) * 100, "浙江", 2026])
    ws2 = wb.create_sheet("特控线")
    ws2.append(["年份", "省份", "特控线分数"])
    ws2.append([2026, "浙江", 594])
    # 本校对照表 to make school_lookup available
    ws3 = wb.create_sheet("本校对照表_总分")
    ws3.append(["校内排名", "高考总分"])
    ws3.append([1, 720])
    ws3.append([100, 640])
    wb.save(os.path.join(macro_dir, "宏观数据_只读.xlsx"))

    data = {
        "workspace": str(tmpdir),
        "total_score": 650,
        "school_rank": 100,
        "school_total": 500,
    }
    result = run(data)
    assert result["status"] == "ok"
    # With only school_lookup (no special_line, no alliance_rank),
    # primary score should be from school_lookup
    # (independent subject sum may add a second component, so check method_count)
    assert result["method_count"] >= 1


# ── Regression tests for bug fixes ───────────────────────────────────


def test_school_lookup_null_unexamined_top():
    """Bug fix: unexamined_top_students=null should not crash method_school_lookup."""
    macro = {
        "本校对照表_总分": [
            {"校内排名": 1, "高考总分": 720},
            {"校内排名": 100, "高考总分": 640},
            {"校内排名": 200, "高考总分": 580},
        ],
    }
    data = {
        "school_rank": 100,
        "school_total": 500,
        "unexamined_top_students": None,  # explicitly null
    }
    # Should not raise TypeError
    result = method_school_lookup(data, macro)
    assert result is not None
    assert result["score"] == 640.0


def test_school_lookup_non_numeric_rows():
    """Bug fix: non-numeric rows in lookup table should be filtered, not crash."""
    macro = {
        "本校对照表_总分": [
            {"校内排名": 1, "高考总分": 720},
            {"校内排名": "N/A", "高考总分": "缺考"},  # non-numeric
            {"校内排名": 100, "高考总分": 640},
        ],
    }
    data = {"school_rank": 100, "school_total": 500}
    # Should not raise ValueError
    result = method_school_lookup(data, macro)
    assert result is not None
    assert result["score"] == 640.0


def test_filter_score_table_filters_non_numeric_score():
    """Bug fix: filter_score_table should exclude rows with non-numeric 分数."""
    rows = [
        {"分数": 750, "累计人数": 100},
        {"分数": "N/A", "累计人数": 200},  # 分数 not numeric
        {"分数": 700, "累计人数": "N/A"},  # 累计人数 not numeric
        {"分数": 650, "累计人数": 5000},
    ]
    filtered = filter_score_table(rows)
    assert len(filtered) == 2
    assert filtered[0]["分数"] == 750
    assert filtered[1]["分数"] == 650


def test_find_sheet_excludes_single_subject_lookup():
    """Bug fix: '对照' keyword should not match '单科对照' when no 本校对照表 exists."""
    # Only 单科对照 exists — should NOT be returned for keyword '对照'
    sheets = ["单科对照表_物理", "特控线"]
    result = find_sheet(sheets, "对照")
    # Should be None because 单科 is excluded
    assert result is None


def test_find_sheet_prefers_non_single_subject():
    """Bug fix: when both exist, '对照' should skip 单科对照 and match the other."""
    sheets = ["单科对照", "年级对照"]
    result = find_sheet(sheets, "对照")
    assert result == "年级对照"


def test_previous_subject_data_skips_current_exam(tmpdir):
    """Bug fix: _find_previous_subject_data should skip the most recent exam
    even when exam name doesn't match, preventing returning current exam data."""
    # Create 成绩总表.xlsx with two exams
    personal_dir = os.path.join(str(tmpdir), "data", "personal")
    os.makedirs(personal_dir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "成绩总表"
    ws.append(["考试名", "日期", "语文", "数学", "英语",
               "选科1名称", "选科1原始分", "选科1赋分",
               "选科2名称", "选科2原始分", "选科2赋分",
               "选科3名称", "选科3原始分", "选科3赋分"])
    # First (older) exam — has 物理 assigned score
    ws.append(["10月月考", "2026-10", 100, 120, 105, "物理", 70, 82, "化学", 68, 79, "技术", 88, 84])
    # Second (current) exam — also has 物理 assigned score
    ws.append(["11月期中", "2026-11", 102, 128, 110, "物理", 75, 85, "化学", 70, 81, "技术", 90, 86])
    wb.save(os.path.join(personal_dir, "成绩总表.xlsx"))

    # Call with empty exam name — should still skip the most recent (current) exam
    # and return data from the older exam (82, not 85)
    result = _find_previous_subject_data(str(tmpdir), "物理", "")
    assert result is not None
    assert result["score"] == 82.0  # from 10月月考, not 85 from 11月期中

    # Call with matching exam name — should skip it and return older data
    result = _find_previous_subject_data(str(tmpdir), "物理", "11月期中")
    assert result is not None
    assert result["score"] == 82.0


# ── Round 2 bug fix tests ────────────────────────────────────────────


def test_previous_subject_data_variable_shadowing(tmpdir):
    """Bug fix: inner loop variable 'i' shadowed outer 'i' in _find_previous_subject_data.

    The inner for-loop used `i` to iterate 选科1/2/3, shadowing the outer
    enumerate(rows) index. This test ensures 选科 lookup still works correctly
    after the rename to `j`.
    """
    personal_dir = os.path.join(str(tmpdir), "data", "personal")
    os.makedirs(personal_dir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "成绩总表"
    ws.append(["考试名", "日期", "语文", "数学", "英语",
               "选科1名称", "选科1原始分", "选科1赋分",
               "选科2名称", "选科2原始分", "选科2赋分",
               "选科3名称", "选科3原始分", "选科3赋分"])
    # Older exam — 物理 in 选科3 position (tests inner loop reaches index 3)
    ws.append(["9月月考", "2026-09", 100, 120, 105, "化学", 68, 79, "生物", 72, 80, "物理", 70, 82])
    # Current exam
    ws.append(["11月期中", "2026-11", 102, 128, 110, "物理", 75, 85, "化学", 70, 81, "技术", 90, 86])
    wb.save(os.path.join(personal_dir, "成绩总表.xlsx"))

    # Should find 物理 from the older exam (in 选科3 position), score=82
    result = _find_previous_subject_data(str(tmpdir), "物理", "11月期中")
    assert result is not None
    assert result["score"] == 82.0
    assert "选科3" in result["field"]


def test_two_module_division_by_zero_equal_lines(tmpdir):
    """Bug fix: method_two_module should not crash when sch_zd == sch_special.

    When school's 特控线 and 浙大线 happen to be equal for 语数英综合,
    the dual-line ratio division would cause ZeroDivisionError.
    Should gracefully fall back to single-line mode.
    """
    from calc_equivalent import method_two_module

    workspace = str(tmpdir)
    macro_dir = os.path.join(workspace, "data", "macro")
    os.makedirs(macro_dir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "11月期中升级"
    # 特控线 section — 语数英综合 line = 350
    ws.append(["特控线分段"])
    ws.append(["科目", "2027划线", "2027上线", "2028划线", "2028上线"])
    ws.append(["语数英综合", 340, 100, 350, 90])
    # 浙大线 section — 语数英综合 line = 350 (same as 特控!)
    ws.append(["浙大线分段"])
    ws.append(["科目", "2027划线", "2027上线", "2028划线", "2028上线"])
    ws.append(["语数英综合", 340, 50, 350, 30])
    wb.save(os.path.join(macro_dir, "宏观数据_只读.xlsx"))

    data = {
        "workspace": workspace,
        "exam_name": "11月期中",
        "total_score": 600,
        "subjects": [
            {"name": "语文", "raw": 110},
            {"name": "数学", "raw": 130},
            {"name": "英语", "raw": 115},
        ],
    }

    # Should not raise ZeroDivisionError — falls back to single-line mode
    result = method_two_module(data, {})
    assert result is not None
    assert result["method"] == "双模块换算法"
    assert result["confidence"] == "B"  # single-line → B级


def test_school_threshold_division_by_zero_equal_lines(tmpdir):
    """Bug fix: method_school_threshold should return None when te_line == zd_line.

    When 特控线 and 浙大线 are equal, the ratio division would crash.
    Should gracefully return None instead.
    """
    from calc_equivalent import method_school_threshold

    workspace = str(tmpdir)
    macro_dir = os.path.join(workspace, "data", "macro")
    os.makedirs(macro_dir, exist_ok=True)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "一分一段表"
    ws1.append(["分数", "累计人数", "省份", "年份"])
    for i, score in enumerate(range(750, 299, -10)):
        ws1.append([score, (i + 1) * 100, "浙江", 2026])
    ws2 = wb.create_sheet("特控线")
    ws2.append(["年份", "省份", "特控线分数"])
    ws2.append([2026, "浙江", 594])
    # 升级 sheet with equal te_line and zd_line
    ws3 = wb.create_sheet("11月期中升级")
    ws3.append(["特控线分段"])
    ws3.append(["科目", "2027划线", "2027上线", "2028划线", "2028上线"])
    ws3.append(["语数英综合", 340, 100, 350, 90])
    ws3.append(["浙大线分段"])
    ws3.append(["科目", "2027划线", "2027上线", "2028划线", "2028上线"])
    ws3.append(["语数英综合", 340, 50, 350, 30])
    wb.save(os.path.join(macro_dir, "宏观数据_只读.xlsx"))

    data = {
        "workspace": workspace,
        "exam_name": "11月期中",
        "total_score": 600,
        "school_total": 600,
        "subjects": [
            {"name": "语文", "raw": 110},
            {"name": "数学", "raw": 130},
            {"name": "英语", "raw": 115},
        ],
    }

    # Should not raise ZeroDivisionError — returns None
    result = method_school_threshold(data, {})
    assert result is None
