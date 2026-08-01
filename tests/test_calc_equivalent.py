"""Test calc_equivalent.py — the core calculation engine."""

import json
import os
import tempfile

import pytest

from calc_equivalent import run


def make_macro_ws(tmpdir):
    """Create minimal macro data for testing."""
    from openpyxl import Workbook
    ws_root = tmpdir
    macro_dir = os.path.join(ws_root, "data", "macro")
    os.makedirs(macro_dir, exist_ok=True)

    wb = Workbook()

    # 一分一段表 sheet
    ws1 = wb.active
    ws1.title = "一分一段表"
    ws1.append(["分数", "累计人数", "省份", "年份"])
    # Simulate: top score 750 = 1 person, each drop by 10 = +1000 people
    for i, score in enumerate(range(750, 299, -10)):
        ws1.append([score, (i + 1) * 100, "浙江", 2026])

    # 特控线 sheet
    ws2 = wb.create_sheet("特控线")
    ws2.append(["年份", "省份", "特控线分数"])
    ws2.append([2026, "浙江", 594])

    # 本校对照表_总分 sheet
    ws3 = wb.create_sheet("本校对照表_总分")
    ws3.append(["校内排名", "高考总分"])
    ws3.append([1, 720])
    ws3.append([10, 700])
    ws3.append([50, 670])
    ws3.append([100, 640])
    ws3.append([200, 600])
    ws3.append([300, 560])

    wb.save(os.path.join(macro_dir, "宏观数据_只读.xlsx"))
    return ws_root


def test_method_score_line(tmpdir):
    """Test 分数线对照法 with valid data."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "special_line_exam": 546.5,
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["primary_method"] == "分数线对照法"
    assert result["confidence"] == "A"
    assert 660 <= result["equivalent_score"] <= 680
    assert result["error_lower"] <= result["equivalent_score"] <= result["error_upper"]


def test_method_percentile(tmpdir):
    """Test 排名锚定法."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "alliance_rank": 2000,
        "alliance_total": 20000,
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["primary_method"] == "排名锚定法"
    assert result["confidence"] == "A"


def test_method_school_lookup(tmpdir):
    """Test 校内排名对照法 (A级)."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "school_rank": 50,
        "school_total": 500,
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["primary_method"] == "校内排名对照法"
    assert result["confidence"] == "A"
    assert abs(result["equivalent_score"] - 670) < 10
    assert result["method_count"] == 1


def make_macro_ws_no_lookup(tmpdir):
    """Create macro data WITHOUT 本校对照表 (so method 4 is the only school method)."""
    from openpyxl import Workbook
    ws_root = tmpdir
    macro_dir = os.path.join(ws_root, "data", "macro")
    os.makedirs(macro_dir, exist_ok=True)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "一分一段表"
    ws1.append(["分数", "累计人数", "省份", "年份"])
    for i, score in enumerate(range(750, 299, -10)):
        ws1.append([score, (i + 1) * 100, "浙江", 2026])

    wb.save(os.path.join(macro_dir, "宏观数据_只读.xlsx"))
    return ws_root


def test_school_rank_no_lookup_insufficient(tmpdir):
    """仅校排名无对照表时，C级被置信度门槛拦截，返回insufficient_data."""
    ws = make_macro_ws_no_lookup(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "school_rank": 80,
        "school_total": 500,
        "school_type": "市重点",
    }
    result = run(data)
    assert result["status"] == "insufficient_data"
    assert "特控线" in result["reason"] or "低精度" in result["reason"]


def make_macro_ws_with_upgrade_and_lookup(tmpdir):
    """Create macro data with upgrade sheet + 本校对照表 for cross-validation tests."""
    from openpyxl import Workbook
    ws_root = tmpdir
    macro_dir = os.path.join(ws_root, "data", "macro")
    os.makedirs(macro_dir, exist_ok=True)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "一分一段表"
    ws1.append(["分数", "累计人数", "省份", "年份"])
    for i, score in enumerate(range(750, 299, -10)):
        ws1.append([score, (i + 1) * 100, "浙江", 2026])

    ws2 = wb.create_sheet("特控线")
    ws2.append(["年份", "省份", "特控线分数"])
    ws2.append([2026, "浙江", 594])

    ws3 = wb.create_sheet("期末升级")
    ws3.append(["科目", "2027划线", "2027上线", "2028划线", "2028上线"])
    ws3.append(["特控分段", "", "", "", ""])
    ws3.append(["语数英综合", "", "", 270, 500])
    ws3.append(["物理", "", "", 65, ""])
    ws3.append(["化学", "", "", 70, ""])
    ws3.append(["浙大分段", "", "", "", ""])
    ws3.append(["语数英综合", "", "", 300, 150])
    ws3.append(["物理", "", "", 85, ""])
    ws3.append(["化学", "", "", 90, ""])

    ws4 = wb.create_sheet("本校对照表_总分")
    ws4.append(["校内排名", "高考总分"])
    ws4.append([1, 720])
    ws4.append([10, 700])
    ws4.append([50, 670])
    ws4.append([100, 640])
    ws4.append([200, 600])
    ws4.append([300, 560])

    wb.save(os.path.join(macro_dir, "宏观数据_只读.xlsx"))
    return ws_root


def test_cross_validation(tmpdir):
    """Test that multiple methods produce cross-validation."""
    ws = make_macro_ws_with_upgrade_and_lookup(tmpdir)
    data = {
        "workspace": ws,
        "exam_name": "期末",
        "total_score": 650,
        "special_line_exam": 546.5,
        "alliance_rank": 3200,
        "alliance_total": 21000,
        "subjects": [
            {"name": "语文", "raw": 120},
            {"name": "数学", "raw": 130},
            {"name": "英语", "raw": 140},
            {"name": "物理", "raw": 80, "assigned": 88},
            {"name": "化学", "raw": 75, "assigned": 85},
        ],
    }
    result = run(data)
    assert result["status"] == "ok"
    # 族① (双模块换算法) 优先级最高，成为主方法；族② (分数线对照法) 作为交叉验证
    assert result["primary_method"] in ("双模块换算法", "分数线对照法")
    assert len(result["cross_validations"]) >= 1
    assert "trust_note" in result


def test_insufficient_data(tmpdir):
    """Test that no data returns proper error."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        # No ranking, no special line
    }
    result = run(data)
    assert result["status"] == "insufficient_data"


def test_no_macro_file(tmpdir):
    """Test that missing macro file returns error."""
    data = {
        "workspace": str(tmpdir),
        "total_score": 650,
    }
    result = run(data)
    assert result["status"] == "error"


def test_priority_order(tmpdir):
    """Test that score_line (P2) beats percentile (P5) when both available."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "special_line_exam": 546.5,
        "alliance_rank": 3200,
        "alliance_total": 21000,
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["primary_method"] == "分数线对照法"  # P1 wins over P3


def test_rank_exceeds_total(tmpdir):
    """Test that rank > total is caught and method returns None."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "alliance_rank": 50000,
        "alliance_total": 1000,  # rank exceeds total
    }
    result = run(data)
    # percentile method should fail (rank > total), no other method has data
    assert result["status"] == "insufficient_data"


def test_percentile_gaoer(tmpdir):
    """Test that 高二 percentile anchoring still gets A-level (grade no longer affects confidence)."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "alliance_rank": 2000,
        "alliance_total": 20000,
        "grade": "高二",
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["primary_method"] == "排名锚定法"
    assert result["confidence"] == "A"


def test_gaoyi_no_longer_blocked(tmpdir):
    """Test that 高一 with ranking data now calculates equivalent score (grade no longer blocks)."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "alliance_rank": 2000,
        "alliance_total": 20000,
        "grade": "高一",
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["confidence"] == "A"


def test_score_line_beats_percentile(tmpdir):
    """Test that score_line (P2) takes priority over percentile (P5)."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 650,
        "special_line_exam": 546.5,
        "alliance_rank": 3200,
        "alliance_total": 21000,
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["primary_method"] == "分数线对照法"
    assert result["confidence"] == "A"


def test_score_scale_450_subject_ratio(tmpdir):
    """B2 fix: 450分制下，单科等效分比例计算应使用原始总分（非换算后的750制）."""
    ws = make_macro_ws(tmpdir)
    data = {
        "workspace": ws,
        "total_score": 314,
        "score_scale": 450,
        "special_line_exam": 286.5,
        "subjects": [
            {"name": "语文", "raw": 98.5},
            {"name": "数学", "raw": 111},
            {"name": "英语", "raw": 104.5},
        ],
    }
    result = run(data)
    assert result["status"] == "ok"
    subject_scores = result.get("subject_scores", [])
    assert len(subject_scores) == 3
    total_subject_sum = sum(s["score"] for s in subject_scores if s["score"])
    # M14: verify ratio uses original 450-scale denominator, not converted 750
    chinese = [s for s in subject_scores if s["subject"] == "语文"][0]
    actual_ratio = chinese["score"] / total_subject_sum
    expected_ratio = 98.5 / 314  # should be ~0.3137, NOT 98.5/523.33≈0.188
    assert abs(actual_ratio - expected_ratio) < 0.01, \
        f"语文占比{actual_ratio:.4f}，期望{expected_ratio:.4f}（B2: 分母应为原始450制，非750制）"


def make_macro_ws_with_upgrade(tmpdir):
    """Create macro data with期末升级 sheet for two-module and school-threshold tests."""
    from openpyxl import Workbook
    ws_root = tmpdir
    macro_dir = os.path.join(ws_root, "data", "macro")
    os.makedirs(macro_dir, exist_ok=True)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "一分一段表"
    ws1.append(["分数", "累计人数", "省份", "年份"])
    for i, score in enumerate(range(750, 299, -10)):
        ws1.append([score, (i + 1) * 100, "浙江", 2026])

    ws2 = wb.create_sheet("特控线")
    ws2.append(["年份", "省份", "特控线分数"])
    ws2.append([2026, "浙江", 594])

    # 期末升级 sheet: col0=科目, col1=2027划线, col2=2027上线, col3=2028划线, col4=2028上线
    ws3 = wb.create_sheet("期末高一下升级")
    ws3.append(["科目", "2027划线", "2027上线", "2028划线", "2028上线"])
    ws3.append(["特控分段", "", "", "", ""])
    ws3.append(["语数英综合", "", "", 270, 500])
    ws3.append(["物理", "", "", 65, ""])
    ws3.append(["化学", "", "", 70, ""])
    ws3.append(["浙大分段", "", "", "", ""])
    ws3.append(["语数英综合", "", "", 300, 150])
    ws3.append(["物理", "", "", 85, ""])
    ws3.append(["化学", "", "", 90, ""])

    # 结构 sheet for school_threshold
    ws4 = wb.create_sheet("期末结构")
    ws4.append(["类别", "人数"])
    ws4.append(["全校总人数", "835人"])

    wb.save(os.path.join(macro_dir, "宏观数据_只读.xlsx"))
    return ws_root


def test_two_module_method(tmpdir):
    """M13: 双模块换算法 (priority 1) with upgrade sheet."""
    ws = make_macro_ws_with_upgrade(tmpdir)
    data = {
        "workspace": ws,
        "exam_name": "高一下期末",
        "total_score": 570,
        "score_scale": 750,
        "subjects": [
            {"name": "语文", "raw": 115},
            {"name": "数学", "raw": 108},
            {"name": "英语", "raw": 112},
            {"name": "物理", "raw": 80, "assigned": 88},
            {"name": "化学", "raw": 75, "assigned": 85},
        ],
    }
    result = run(data)
    assert result["status"] == "ok"
    assert result["primary_method"] == "双模块换算法"
    assert result["confidence"] in ("A", "B")


def test_school_threshold_method(tmpdir):
    """M13: 校排阈值估算法 (priority 3) triggered with upgrade + school data."""
    ws = make_macro_ws_with_upgrade(tmpdir)
    data = {
        "workspace": ws,
        "exam_name": "高一下期末",
        "total_score": 480,  # → 288 in 450-scale, between 特控270 and 浙大300
        "score_scale": 750,
        "school_type": "省重点",
        "subjects": [
            {"name": "语文", "raw": 105},
            {"name": "数学", "raw": 100},
            {"name": "英语", "raw": 95},
            {"name": "物理", "raw": 60},
            {"name": "化学", "raw": 55},
        ],
    }
    result = run(data)
    assert result["status"] == "ok"
    # 双模块 should be primary (P1), 校排阈值 should appear as cross-validation
    methods_seen = [m["method"] for m in result["method_details"]]
    assert "校排阈值估算法" in methods_seen





# ── 共享工具函数测试 ──

def test_safe_float():
    from calc_equivalent import safe_float
    assert safe_float("123.5") == 123.5
    assert safe_float(42) == 42.0
    assert safe_float("abc") is None
    assert safe_float(None) is None
    assert safe_float(0) == 0.0


def test_find_latest_gaokao_special_line():
    from calc_equivalent import find_latest_gaokao_special_line
    lines = [
        {"年份": 2024, "特控线分数": 590},
        {"年份": 2026, "特控线分数": 594},
        {"年份": 2025, "特控线分数": 592},
    ]
    score, year = find_latest_gaokao_special_line(lines)
    assert score == 594.0
    assert year == 2026


def test_find_latest_gaokao_special_line_empty():
    from calc_equivalent import find_latest_gaokao_special_line
    score, year = find_latest_gaokao_special_line([])
    assert score is None
    assert year is None


def test_find_latest_gaokao_special_line_invalid_year():
    from calc_equivalent import find_latest_gaokao_special_line
    lines = [
        {"年份": "abc", "特控线分数": 590},
        {"年份": 2026, "特控线分数": 594},
    ]
    score, year = find_latest_gaokao_special_line(lines)
    assert score == 594.0
    assert year == 2026


def test_compute_main_raw_sum_all_present():
    from calc_equivalent import compute_main_raw_sum
    data = {
        "total_score": 650,
        "subjects": [
            {"name": "语文", "raw": 120},
            {"name": "数学", "raw": 130},
            {"name": "英语", "raw": 110},
        ],
    }
    result = compute_main_raw_sum(data)
    assert result == 360.0


def test_compute_main_raw_sum_partial():
    from calc_equivalent import compute_main_raw_sum
    data = {
        "total_score": 650,
        "subjects": [
            {"name": "语文", "raw": 120},
            {"name": "数学", "raw": 130},
        ],
    }
    # Only 2 of 3 main subjects → proportional fallback
    result = compute_main_raw_sum(data)
    assert result > 0


def test_find_score_by_count_basic():
    from calc_equivalent import find_score_by_count
    table = [
        {"分数": 700, "累计人数": 100},
        {"分数": 650, "累计人数": 500},
        {"分数": 600, "累计人数": 1000},
    ]
    result = find_score_by_count(table, 500)
    assert result == 650.0


def test_find_score_by_count_empty():
    from calc_equivalent import find_score_by_count
    result = find_score_by_count([], 500)
    assert result is None


def test_find_score_by_count_invalid_rows():
    from calc_equivalent import find_score_by_count
    table = [
        {"分数": "abc", "累计人数": 100},
        {"分数": 650, "累计人数": 500},
    ]
    result = find_score_by_count(table, 500)
    assert result == 650.0


def test_run_missing_total_score():
    """P0: 缺少 total_score 应返回明确错误"""
    data = {"workspace": "/tmp"}
    result = run(data)
    assert result["status"] == "error"
    assert "total_score" in result["reason"]


def test_run_invalid_total_score():
    """P0: total_score 非数值应返回明确错误"""
    data = {"workspace": "/tmp", "total_score": "abc"}
    result = run(data)
    assert result["status"] == "error"


def test_method_percentile_negative_total():
    """P1: total 为负数时应返回 None"""
    from calc_equivalent import method_percentile
    data = {"alliance_rank": 100, "alliance_total": -1}
    result = method_percentile(data, {"一分一段表": []})
    assert result is None


# ── P2/P3 修复测试 ──


def test_find_latest_gaokao_special_line_missing_key():
    """P2: 特控线分数字段缺失时不应抛出 KeyError"""
    from calc_equivalent import find_latest_gaokao_special_line
    lines = [
        {"年份": 2024, "特控线分数": 590},
        {"年份": 2026},  # 缺少特控线分数
        {"年份": 2025, "特控线分数": "abc"},  # 非数值
    ]
    score, year = find_latest_gaokao_special_line(lines)
    assert score == 590.0
    assert year == 2024  # 只有 2024 有有效分数


def test_parse_upgrade_sheet_returns_rank_data():
    """P2: parse_upgrade_sheet 应同时返回分数线和上线人数"""
    from calc_equivalent import parse_upgrade_sheet
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "期末升级"
    ws.append(["科目", "2027划线", "2027上线", "2028划线", "2028上线"])
    ws.append(["特控分段", "", "", "", ""])
    ws.append(["语数英综合", "", "", 270, 500])
    ws.append(["物理", "", "", 65, 300])
    ws.append(["浙大分段", "", "", "", ""])
    ws.append(["语数英综合", "", "", 300, 150])
    ws.append(["物理", "", "", 85, 50])

    result = parse_upgrade_sheet(ws)

    # 验证分数线
    assert result["语数英综合"]["special"] == 270
    assert result["语数英综合"]["zd"] == 300
    assert result["物理"]["special"] == 65
    assert result["物理"]["zd"] == 85

    # 验证上线人数
    assert result["语数英综合"]["special_rank"] == 500
    assert result["语数英综合"]["zd_rank"] == 150
    assert result["物理"]["special_rank"] == 300
    assert result["物理"]["zd_rank"] == 50


def test_compute_subject_equivalents_invalid_raw():
    """P2: 语数英原始分无法转为 float 时不应 NameError，应按均分估算"""
    from calc_equivalent import compute_subject_equivalents
    data = {
        "total_score": 650,
        "_total_equivalent": 600,
        "subjects": [
            {"name": "语文", "raw": "缺考"},  # 非数值
            {"name": "数学", "raw": 120},
            {"name": "英语", "raw": 110},
            {"name": "物理", "assigned": 88},
        ],
    }
    # 不应抛出 NameError，语文应有估算分数（均分估算法，C级）
    results = compute_subject_equivalents(data, {})
    chinese = [r for r in results if r["subject"] == "语文"]
    assert len(chinese) == 1
    assert chinese[0]["score"] is not None  # 按均分估算
    assert chinese[0]["confidence"] == "C"
    assert chinese[0]["method"] == "均分估算法"
    # 数学、英语应有有效分数
    math = [r for r in results if r["subject"] == "数学"][0]
    english = [r for r in results if r["subject"] == "英语"][0]
    assert math["score"] is not None
    assert english["score"] is not None


# ── P0: workspace 未定义 bug 回归测试 ──


def test_two_module_workspace_undefined_fallback(tmpdir):
    """P0 回归: method_two_module 中选科无校内划线+无赋分时，
    回退至 _find_previous_subject_data 不应因 workspace 未定义而 NameError。

    重现条件：
      1. 升级 Sheet 存在（触发 method_two_module）
      2. 某选科在升级 Sheet 中无分数线（走 Priority 3 回退）
      3. 该选科无赋分（走 _find_previous_subject_data 跨次回退）
      4. 成绩总表中有历史数据可供回退
    """
    from calc_equivalent import method_two_module
    from openpyxl import Workbook

    workspace = str(tmpdir)
    macro_dir = os.path.join(workspace, "data", "macro")
    os.makedirs(macro_dir, exist_ok=True)

    # 创建升级 Sheet（有语数英综合划线，但无物理划线）
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "期末升级"
    ws1.append(["特控线分段"])
    ws1.append(["科目", "2027划线", "2027上线", "2028划线", "2028上线"])
    ws1.append(["语数英综合", "", "", 270, 500])
    # 物理在升级 Sheet 中没有划线数据 → 触发 Priority 3 回退
    ws1.append(["浙大线分段"])
    ws1.append(["科目", "2027划线", "2027上线", "2028划线", "2028上线"])
    ws1.append(["语数英综合", "", "", 300, 150])
    wb.save(os.path.join(macro_dir, "宏观数据_只读.xlsx"))

    # 创建成绩总表，含历史物理赋分数据
    personal_dir = os.path.join(workspace, "data", "personal")
    os.makedirs(personal_dir, exist_ok=True)
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "成绩总表"
    ws2.append(["考试名", "日期", "语文", "数学", "英语",
                "选科1名称", "选科1原始分", "选科1赋分",
                "选科2名称", "选科2原始分", "选科2赋分",
                "选科3名称", "选科3原始分", "选科3赋分"])
    ws2.append(["10月月考", "2026-10", 100, 120, 105,
                "物理", 70, 82, "化学", 68, 79, "技术", 88, 84])
    wb2.save(os.path.join(personal_dir, "成绩总表.xlsx"))

    data = {
        "workspace": workspace,
        "exam_name": "期末",
        "total_score": 570,
        "subjects": [
            {"name": "语文", "raw": 115},
            {"name": "数学", "raw": 108},
            {"name": "英语", "raw": 112},
            # 物理无赋分，升级 Sheet 中也无物理划线 → 触发 _find_previous_subject_data
            {"name": "物理", "raw": 80},
        ],
    }

    # 修复前：NameError: name 'workspace' is not defined
    # 修复后：应正常返回结果
    result = method_two_module(data, {})
    assert result is not None
    assert result["method"] == "双模块换算法"
    # 物理应通过跨次回退获得分数（C级）
    assert "物理" in result["detail"]


def test_normalize_confidence_various_inputs():
    """P2: normalize_confidence 应正确处理各种输入格式"""
    from calc_equivalent import normalize_confidence
    assert normalize_confidence("A") == "A"
    assert normalize_confidence("A级") == "A"
    assert normalize_confidence("a") == "A"
    assert normalize_confidence("B级") == "B"
    assert normalize_confidence(None) == "B"
    assert normalize_confidence("") == "B"
    assert normalize_confidence("X") == "B"  # 未知值默认 B
    assert normalize_confidence("D级") == "D"
