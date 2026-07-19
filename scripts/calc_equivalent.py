#!/usr/bin/env python3
"""Calculate equivalent Gaokao score using all available methods.

Priority order:
  1. 分数线对照法 (Score-line comparison) — A级
  2. 校内排名对照法 (School ranking lookup) — A级
  3. 全市/联盟排名锚定法 (Percentile anchoring) — A级
  4. 校排名估算 (School rank estimation) — C级

Input: JSON via stdin
Output: JSON with primary method, score, confidence, error range, cross-validations
"""

import json
import os
import sys

from openpyxl import load_workbook


def read_sheet_rows(ws):
    """Read all rows from a worksheet as list of dicts (header row is row 1)."""
    if ws.max_row < 2:
        return []
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


def read_macro_data(workspace):
    """Read all macro data sheets."""
    path = os.path.join(workspace, "数据", "宏观数据", "宏观数据_只读.xlsx")
    if not os.path.exists(path):
        return None
    wb = load_workbook(path, data_only=True)
    data = {}
    for name in wb.sheetnames:
        data[name] = read_sheet_rows(wb[name])
    return data


def method_score_line(data, macro):
    """Method 1: 分数线对照法（等比例放缩）— A级."""
    special_line_exam = data.get("special_line_exam")
    if not special_line_exam:
        return None

    special_lines = macro.get("特控线", [])
    if not special_lines:
        return None

    # Find the most recent gaokao special line
    gaokao_sl = None
    for sl in special_lines:
        if sl.get("特控线分数"):
            gaokao_sl = float(sl["特控线分数"])
            break

    if not gaokao_sl:
        return None

    total_score = float(data["total_score"])
    if total_score == 750:
        return {
            "method": "分数线对照法",
            "score": 750.0,
            "confidence": "A",
            "detail": f"满分750 → 等效分750",
        }

    es = (750 - gaokao_sl) / (750 - special_line_exam) * (total_score - special_line_exam) + gaokao_sl
    return {
        "method": "分数线对照法",
        "score": round(es, 1),
        "confidence": "A",
        "detail": f"等效分 = (750-{gaokao_sl})/(750-{special_line_exam})×({total_score}-{special_line_exam})+{gaokao_sl} = {es:.1f}",
    }


def method_school_lookup(data, macro):
    """Method 2: 校内排名对照法 — A级."""
    school_rank = data.get("school_rank")
    if not school_rank:
        return None

    school_total = data.get("school_total")
    class_type = data.get("class_type")
    unexamined_top = data.get("unexamined_top_students", 0)

    # Apply class-type calibration
    calibrated_rank = int(school_rank) + int(unexamined_top)

    lookup_sheet = macro.get("本校对照表_总分", [])
    if not lookup_sheet:
        return None

    # Sort by 校内排名
    lookup = sorted(lookup_sheet, key=lambda r: int(r.get("校内排名", 0)))
    if not lookup:
        return None

    ranks = [int(r["校内排名"]) for r in lookup]
    scores = [float(r["高考总分"]) for r in lookup]

    # Exact match
    if calibrated_rank in ranks:
        idx = ranks.index(calibrated_rank)
        score = scores[idx]
    elif calibrated_rank < ranks[0]:
        score = scores[0]
    elif calibrated_rank > ranks[-1]:
        score = scores[-1]
    else:
        # Linear interpolation
        for i in range(len(ranks) - 1):
            if ranks[i] <= calibrated_rank <= ranks[i + 1]:
                ratio = (calibrated_rank - ranks[i]) / (ranks[i + 1] - ranks[i])
                score = scores[i] + ratio * (scores[i + 1] - scores[i])
                break
        else:
            return None

    detail = f"校内排名{calibrated_rank}名"
    if unexamined_top:
        detail += f"（原始排名{school_rank}，补算重点班{unexamined_top}人）"
    detail += f" → 对照表对应高考总分{score:.0f}分"

    return {
        "method": "校内排名对照法",
        "score": round(score, 1),
        "confidence": "A",
        "detail": detail,
    }


def method_percentile(data, macro):
    """Method 3: 全市/联盟排名锚定法 — A级."""
    rank = data.get("city_rank") or data.get("alliance_rank")
    total = data.get("city_total") or data.get("alliance_total")

    if not rank or not total:
        return None

    rank = int(rank)
    total = int(total)

    if rank > total:
        return None

    percentile = 1.0 - (rank / total)
    score_table = macro.get("一分一段表", [])
    if not score_table:
        return None

    # Sort by 分数 descending (higher score = lower cumulative count)
    sorted_table = sorted(score_table, key=lambda r: int(r.get("分数", 0)), reverse=True)

    # Find the score where cumulative_count/total ≈ percentile
    # 一分一段表: 分数 -> 累计人数. 累计人数/总人数 ≈ percentile
    # We need the total number of test takers first
    max_count = max(int(r.get("累计人数", 0)) for r in sorted_table)
    if max_count == 0:
        return None

    target_count = int((1 - percentile) * max_count)

    # Find closest match
    best_score = None
    best_diff = float("inf")
    for row in sorted_table:
        count = int(row.get("累计人数", 0))
        diff = abs(count - target_count)
        if diff < best_diff:
            best_diff = diff
            best_score = float(row["分数"])

    if best_score is None:
        return None

    source = "全市排名" if data.get("city_rank") else "联盟排名"
    return {
        "method": "排名锚定法",
        "score": round(best_score, 1),
        "confidence": "A",
        "detail": f"{source}{rank}/{total} → 百分位{percentile:.3f} → 等效分{best_score:.0f}",
    }


def method_school_estimate(data, macro):
    """Method 4: 校排名估算 — C级."""
    school_rank = data.get("school_rank")
    if not school_rank:
        return None

    school_type = data.get("school_type", "市重点")
    coefficients = {"省重点": 15, "市重点": 8, "普通高中": 3}
    coef = coefficients.get(school_type, 8)

    estimated_city_rank = int(school_rank) * coef

    score_table = macro.get("一分一段表", [])
    if not score_table:
        return None

    sorted_table = sorted(score_table, key=lambda r: int(r.get("分数", 0)), reverse=True)
    max_count = max(int(r.get("累计人数", 0)) for r in sorted_table)
    if max_count == 0:
        return None

    # Rough estimate: use the estimated city rank to find score
    target_count = estimated_city_rank

    best_score = None
    best_diff = float("inf")
    for row in sorted_table:
        count = int(row.get("累计人数", 0))
        diff = abs(count - target_count)
        if diff < best_diff:
            best_diff = diff
            best_score = float(row["分数"])

    if best_score is None:
        return None

    return {
        "method": "校排名估算",
        "score": round(best_score, 1),
        "confidence": "C",
        "detail": f"校内排名{school_rank} × {school_type}系数{coef} ≈ 全市{estimated_city_rank}名 → 等效分{best_score:.0f}",
    }


def compute_error_range(primary, cross_validations):
    """Calculate error range based on cross-validation spread."""
    if not cross_validations:
        default_margin = {"A": 5, "B": 8, "C": 15, "D": 20}
        margin = default_margin.get(primary["confidence"], 10)
        return {
            "lower": round(primary["score"] - margin, 1),
            "upper": round(primary["score"] + margin, 1),
        }

    valid_cv = [cv for cv in cross_validations if cv.get("score")]
    if not valid_cv:
        margin = 5
    else:
        avg_cv = sum(cv["score"] for cv in valid_cv) / len(valid_cv)
        margin = max(3, abs(primary["score"] - avg_cv))

    return {
        "lower": round(primary["score"] - margin, 1),
        "upper": round(primary["score"] + margin, 1),
    }


def run(data):
    workspace = os.path.abspath(data.get("workspace", "."))
    macro = read_macro_data(workspace)

    if macro is None:
        return {
            "status": "error",
            "reason": "宏观数据_只读.xlsx 不存在，请先完成初始设置",
        }

    methods = []

    # Try all methods in priority order
    result = method_score_line(data, macro)
    if result:
        methods.append(result)

    result = method_school_lookup(data, macro)
    if result:
        methods.append(result)

    result = method_percentile(data, macro)
    if result:
        methods.append(result)

    result = method_school_estimate(data, macro)
    if result:
        methods.append(result)

    if not methods:
        return {
            "status": "insufficient_data",
            "reason": "当前数据不足以计算等效分。至少需要以下之一：模考特控线、校内排名+对照表、市/联盟排名+一分一段表。",
        }

    primary = methods[0]  # Highest priority
    cross_validations = []
    for m in methods[1:]:
        diff = round(m["score"] - primary["score"], 1)
        cross_validations.append({
            "method": m["method"],
            "score": m["score"],
            "confidence": m["confidence"],
            "difference": diff,
        })

    error = compute_error_range(primary, cross_validations)

    trust_note = None
    if cross_validations:
        diffs = [abs(cv["difference"]) for cv in cross_validations]
        if max(diffs) <= 3:
            trust_note = "交叉验证差异在±3分以内，等效分可信度较高"

    return {
        "status": "ok",
        "primary_method": primary["method"],
        "equivalent_score": primary["score"],
        "confidence": primary["confidence"],
        "error_lower": error["lower"],
        "error_upper": error["upper"],
        "calculation_detail": primary.get("detail", ""),
        "cross_validations": cross_validations,
        "trust_note": trust_note,
        "warnings": [],
    }


def main():
    data = json.loads(sys.stdin.read())
    result = run(data)
    print(json.dumps(result, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
